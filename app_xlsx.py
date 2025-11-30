import io
import os
import re
import json
import time
import glob
import shutil
from datetime import datetime
from typing import Optional, List, Tuple, Dict
import pandas as pd
import streamlit as st
from PIL.ImagePalette import raw
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------------------------------
# Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ø£Ø³Ø§Ø³ÙŠØ© (Ù…Ø³Ø§Ø± Ø«Ø§Ø¨Øª ÙƒÙ…Ø§ Ø·Ù„Ø¨Øª)
# -------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

EXCEL_PATH = os.path.join(DATA_DIR, "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†.xlsx")
CONFIG_PATH = os.path.join(DATA_DIR, "config.json")

# Ø­Ù‚ÙˆÙ‚ Ø§Ù„Ù…Ø·ÙˆØ±
DEV_NAME = "sevens"

# Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§ÙØªØ±Ø§Ø¶ÙŠØ©
DEFAULT_CONFIG = {
    "global_min_level": 2,
    "enable_backups": False,
    "backup_keep": 0,
    "code_case": "upper",
    "auto_suffix_mode": "by_checkbox",
    "enforce_suffix": False,
    "suffix_text": "-S",
    "suffix_apply_on": ["scan", "bulk", "merge", "ops", "editor", "import"],
    "suffix_apply_on_contexts": ["scan", "bulk", "ops", "stocktake", "add"],
}

SCAN_HISTORY_MAX = 500

# -------------------------------------------------
# ÙˆØ§Ø¬Ù‡Ø© ÙˆØªÙ†Ø³ÙŠÙ‚
# -------------------------------------------------
st.set_page_config(page_title="Ù†Ø¸Ø§Ù… Ù…Ø®Ø²ÙˆÙ† Ù‚Ø·Ø¹ Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª (Excel)", layout="wide")
st.markdown(
    """
   <style>
/* =========================================================
ğŸŒŒ SEVENS NEXT Dashboard â€” ØªØµÙ…ÙŠÙ… Ø§Ø­ØªØ±Ø§ÙÙŠ ÙØ§Ø®Ø±
Ø¥ØµØ¯Ø§Ø± 2025 â€” Ø£Ø³Ù„ÙˆØ¨ Ø£Ø²Ø±Ù‚ Ø³Ù…Ø§ÙˆÙŠ Ø£Ù†ÙŠÙ‚ Ø¨Ø®Ø· Tajawal
========================================================= */

body {
  direction: rtl;
  text-align: right;
  font-family: 'Tajawal', sans-serif !important;
  background: linear-gradient(135deg, #f7faff 0%, #eef5fb 100%);
  color: #1f2d3d;
  margin: 0;
  padding: 0;
}

/* ğŸ¯ Ø§Ù„Ø´Ø±ÙŠØ· Ø§Ù„Ø¬Ø§Ù†Ø¨ÙŠ */
[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #0052cc 0%, #00bcd4 100%) !important;
  color: white !important;
  box-shadow: 3px 0 20px rgba(0, 0, 0, 0.15);
}
[data-testid="stSidebar"] * {
  color: #fff !important;
  direction: rtl;
  text-align: right;
  font-size: 15px;
}
[data-testid="stSidebar"] .sidebar-content {
  padding-top: 20px !important;
}

/* ğŸ§­ Ø¹Ù†Ø§ÙˆÙŠÙ† Ø§Ù„Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø¬Ø§Ù†Ø¨ÙŠØ© */
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
  color: #fff !important;
  font-weight: 700;
  text-shadow: 0 2px 5px rgba(0, 0, 0, 0.25);
}

/* ğŸ“¦ Ø¨Ø·Ø§Ù‚Ø§Øª Ø§Ù„Ù…Ø¤Ø´Ø±Ø§Øª */
.metric-box {
  background: linear-gradient(145deg, #ffffff, #f2f6fc);
  border-radius: 18px;
  padding: 25px;
  text-align: center;
  box-shadow: 0 6px 16px rgba(0, 0, 0, 0.06);
  transition: all 0.3s ease;
  border: 1px solid #e3ebf5;
}
.metric-box:hover {
  transform: translateY(-5px);
  box-shadow: 0 10px 25px rgba(0, 0, 0, 0.08);
}
.metric-box h3 {
  color: #007bff;
  margin: 8px 0;
  font-size: 30px;
  font-weight: 800;
}
.metric-box p {
  color: #666;
  font-size: 15px;
  margin: 0;
}

/* ğŸ§© Ø§Ù„Ø¨Ø·Ø§Ù‚Ø§Øª Ø§Ù„Ø¹Ø§Ù…Ø© */
.card {
  background: white;
  border-radius: 20px;
  padding: 28px;
  box-shadow: 0 5px 20px rgba(0, 0, 0, 0.05);
  margin-bottom: 25px;
  border: 1px solid #e9eef5;
  transition: all 0.3s ease;
}
.card:hover {
  box-shadow: 0 10px 25px rgba(0, 0, 0, 0.08);
  transform: translateY(-3px);
}

/* ğŸ’ Ø£Ø²Ø±Ø§Ø± SEVENS */
.stButton>button, .btn-main {
  background: linear-gradient(90deg, #007bff 0%, #00bcd4 100%);
  color: white !important;
  border: none;
  padding: 10px 28px;
  border-radius: 12px;
  font-weight: 700;
  font-size: 15px;
  letter-spacing: 0.3px;
  transition: all 0.25s;
  box-shadow: 0 4px 12px rgba(0, 123, 255, 0.25);
}
.stButton>button:hover, .btn-main:hover {
  background: linear-gradient(90deg, #00bcd4 0%, #007bff 100%);
  transform: translateY(-2px);
  box-shadow: 0 6px 18px rgba(0, 123, 255, 0.35);
}

/* ğŸ§¾ Ø§Ù„Ø¬Ø¯Ø§ÙˆÙ„ */
table {
  border-collapse: collapse !important;
  border-radius: 12px;
  overflow: hidden;
}
th {
  background: linear-gradient(90deg, #007bff 0%, #00bcd4 100%) !important;
  color: white !important;
  font-weight: 600;
  font-size: 14px;
  text-align: center !important;
  border: none !important;
}
td {
  text-align: center !important;
  padding: 8px 10px !important;
  border: none !important;
}
tbody tr:nth-child(even) {
  background-color: #f9fbff !important;
}
tbody tr:hover {
  background: #eaf4ff !important;
  transition: 0.25s;
}

/* âœ… ØµÙ†Ø§Ø¯ÙŠÙ‚ Ø§Ù„Ø­Ø§Ù„Ø© */
.success-box {
  background: #ecfff6;
  border: 1px solid #a8f5d0;
  color: #0f5132;
  padding: 12px 18px;
  border-radius: 14px;
}
.warn-box {
  background: #fff9e6;
  border: 1px solid #ffe680;
  color: #946200;
  padding: 12px 18px;
  border-radius: 14px;
}
.error-box {
  background: #fff2f2;
  border: 1px solid #ffb3b3;
  color: #991b1b;
  padding: 12px 18px;
  border-radius: 14px;
}

/* ğŸ’  Ø§Ù„Ø¹Ù†Ø§ÙˆÙŠÙ† */
h1, h2, h3, h4 {
  color: #1b2734;
  font-weight: 800;
  letter-spacing: -0.2px;
}
h1 {
  font-size: 28px;
}
h2 {
  font-size: 22px;
}

/* âš™ï¸ Ø§Ù„Ø¥Ø¯Ø®Ø§Ù„Ø§Øª */
input, select, textarea {
  border: 1px solid #cfd8e3 !important;
  border-radius: 12px !important;
  padding: 8px 12px !important;
  font-family: 'Tajawal', sans-serif !important;
  background-color: #fff;
}
input:focus, select:focus, textarea:focus {
  border-color: #00aaff !important;
  box-shadow: 0 0 6px rgba(0, 123, 255, 0.3);
  outline: none !important;
}

/* ğŸŒŸ Ø´Ø¹Ø§Ø± SEVENS */
.logo-box {
  display: flex;
  align-items: center;
  justify-content: center;
  gap: 14px;
  margin-bottom: 25px;
}
.logo-box img {
  height: 48px;
  filter: drop-shadow(0 3px 4px rgba(0,0,0,0.2));
}
.logo-box h1 {
  font-size: 23px;
  font-weight: 800;
  color: #ffffff;
  text-shadow: 0 2px 6px rgba(0,0,0,0.2);
}

/* ğŸ§  Ø­Ù‚ÙˆÙ‚ Ø§Ù„Ù…Ø·ÙˆØ± */
.dev-credit {
  position: fixed;
  bottom: 12px;
  right: 20px;
  background: rgba(0, 123, 255, 0.08);
  padding: 8px 14px;
  border-radius: 10px;
  font-size: 12px;
  color: #007bff;
  z-index: 999;
  backdrop-filter: blur(8px);
}

/* ğŸ“± ØªØ¬Ø§ÙˆØ¨ ÙƒØ§Ù…Ù„ Ù„Ù„Ø¬ÙˆØ§Ù„ */
@media (max-width: 768px) {
  .metric-box h3 { font-size: 22px; }
  .metric-box p { font-size: 13px; }
  .card { padding: 18px; }
  .stButton>button { width: 100%; }
  h1 { font-size: 22px; }
}
</style>


    """,
    unsafe_allow_html=True,
)


# -------------------------------------------------
# Helpers Ø¹Ø§Ù…Ø©
# -------------------------------------------------
def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _file_mtime(path: str) -> float:
    try:
        return os.path.getmtime(path)
    except Exception:
        return 0.0


def _safe_int(x, default=0):
    try:
        return int(float(x))
    except Exception:
        return default


def _unique_order(seq: List[str]) -> List[str]:
    return list(dict.fromkeys(seq))


# -------------------------------------------------
# Ø¥Ø¹Ø¯Ø§Ø¯/Ù‚Ø±Ø§Ø¡Ø© Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª
# -------------------------------------------------
def load_config() -> dict:
    try:
        if os.path.exists(CONFIG_PATH):
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            for k, v in DEFAULT_CONFIG.items():
                cfg.setdefault(k, v)
            cfg["enable_backups"] = False
            cfg["backup_keep"] = 0
            return cfg
    except Exception:
        pass
    return DEFAULT_CONFIG.copy()


def save_config(cfg: dict):
    try:
        cfg["enable_backups"] = False
        cfg["backup_keep"] = 0
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# -------------------------------------------------
# Ù‚ÙÙ„ ÙƒØªØ§Ø¨Ø© Ø¨Ø³ÙŠØ· + ÙƒØªØ§Ø¨Ø© Ø°Ø±Ù‘ÙŠØ© (Ø¨Ø¯ÙˆÙ† Ù†Ø³Ø® Ø§Ø­ØªÙŠØ§Ø·ÙŠ)
# -------------------------------------------------
class SimpleFileLock:
    def __init__(self, target: str, timeout: float = 5.0, interval: float = 0.1):
        self.lock_path = target + ".lock"
        self.timeout = timeout
        self.interval = interval

    def __enter__(self):
        start = time.time()
        while True:
            try:
                fd = os.open(self.lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                os.close(fd)
                break
            except FileExistsError:
                if time.time() - start > self.timeout:
                    break
                time.sleep(self.interval)

    def __exit__(self, exc_type, exc, tb):
        try:
            if os.path.exists(self.lock_path):
                os.remove(self.lock_path)
        except Exception:
            pass


def _atomic_write_excel(writer_fn, dst_path: str):
    tmp_path = dst_path + ".__tmp__.xlsx"
    writer_fn(tmp_path)
    os.replace(tmp_path, dst_path)


def _backup_if_needed():
    return


def write_all_with_retry(stock: pd.DataFrame, minlvl_unused: pd.DataFrame, tx: pd.DataFrame,
                         retries: int = 3, sleep_s: float = 0.6):
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            write_all(stock, minlvl_unused, tx)
            return
        except Exception as e:
            last_err = e
            time.sleep(sleep_s)
    raise last_err


# -------------------------------------------------
# ØªÙ‡ÙŠØ¦Ø© Ù…Ù„Ù Ø§Ù„Ø¥ÙƒØ³Ù„
# -------------------------------------------------
def ensure_excel_file():
    if os.path.exists(EXCEL_PATH):
        return
    stock = pd.DataFrame(columns=["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ù…ÙˆÙ‚Ø¹", "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"])
    tx = pd.DataFrame(
        columns=["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„Ù†ÙˆØ¹", "Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ù…Ù†_Ù…ÙˆÙ‚Ø¹", "Ø§Ù„Ù‰_Ù…ÙˆÙ‚Ø¹", "Ø§Ù„ÙƒÙ…ÙŠØ©", "Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…", "Ù…Ù„Ø§Ø­Ø¸Ø©"])

    def _write(p):
        with pd.ExcelWriter(p, engine="openpyxl", mode="w") as w:
            stock.to_excel(w, index=False, sheet_name="Stock")
            tx.to_excel(w, index=False, sheet_name="Transactions")

    with SimpleFileLock(EXCEL_PATH):
        _atomic_write_excel(_write, EXCEL_PATH)


def _drop_sheet_if_exists(path: str, sheet_name: str):
    try:
        if not os.path.exists(path):
            return
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
            wb.save(path)
    except Exception:
        pass


# -------------------------------------------------
# Ù…Ù†Ø·Ù‚ Ù„Ø§Ø­Ù‚Ø© Ø§Ù„Ø£ØµÙ„ÙŠ + Ù…ÙØ­Ø³Ù‘Ù†Ø§Øª Ø§Ù„Ù…Ø³Ø­ (ØªÙ… ØªØ¹Ø¯ÙŠÙ„Ù‡!)
# -------------------------------------------------
CODE_IN_BRACKETS = re.compile(r"\[([^\[\]]+)\]")
CODE_TOKEN = re.compile(r"[0-9A-Za-z\u0600-\u06FF\-_.\/]+")
_AR_NUM_MAP = str.maketrans("Ù Ù¡Ù¢Ù£Ù¤Ù¥Ù¦Ù§Ù¨Ù©", "0123456789")


def _to_ascii_digits(s: str) -> str:
    return (s or "").translate(_AR_NUM_MAP)


def _sanitize_code_input(text: str) -> str:
    s = "" if text is None else str(text)
    s = _to_ascii_digits(s)
    m = CODE_IN_BRACKETS.search(s)
    if m:
        s = m.group(1)
    s = re.sub(r"[^0-9A-Za-z\u0600-\u06FF\-_.\/]", "", s)
    return s.strip()


def _suffix_to_use(cfg: dict) -> str:
    s = str(cfg.get("suffix_text", "-S"))
    cc = cfg.get("code_case", "upper")
    if cc == "upper":
        return s.upper()
    if cc == "lower":
        return s.lower()
    return s


def _normalize_code_text(text: str, cfg: dict, context: str = "") -> str:
    s = ("" if text is None else str(text)).strip()
    s = _to_ascii_digits(s)
    cc = cfg.get("code_case", "upper")
    if cc == "upper":
        s = s.upper()
    elif cc == "lower":
        s = s.lower()
    return s


def _extract_code_from_text(text: str) -> Optional[str]:
    s = _sanitize_code_input(text)
    if not s:
        return None
    m = CODE_TOKEN.search(s)
    return m.group(0).strip() if m else None


# âœ… ØªÙ… ØªØ¹Ø¯ÙŠÙ„ Ù‡Ø°Ø§ Ø§Ù„Ù…Ù†Ø·Ù‚ Ø¨Ø§Ù„ÙƒØ§Ù…Ù„
def is_original_code(code: str, cfg: dict) -> bool:
    """Ø§Ù„ÙƒÙˆØ¯ Ø§Ù„Ø£ØµÙ„ÙŠ Ù‡Ùˆ Ø§Ù„Ø°ÙŠ Ù„Ø§ ÙŠØ­ØªÙˆÙŠ Ø¹Ù„Ù‰ -S ÙÙŠ Ø§Ù„Ù†Ù‡Ø§ÙŠØ©."""
    suf = _suffix_to_use(cfg)
    return not str(code or "").strip().endswith(suf)


def ensure_original_flag(code: str, cfg: dict, want_original: bool) -> str:
    """Ø¥Ø°Ø§ Ø£Ø±Ø¯Ù†Ø§ ÙƒÙˆØ¯Ù‹Ø§ Ø£ØµÙ„ÙŠÙ‹Ø§ØŒ Ù†Ø²ÙŠÙ„ -S. Ø¥Ø°Ø§ Ø£Ø±Ø¯Ù†Ø§ ØªÙ‚Ù„ÙŠØ¯Ù‹Ø§ØŒ Ù†Ø¶ÙŠÙ -S."""
    c = (code or "").strip()
    suf = _suffix_to_use(cfg)
    if want_original:
        # Ù†Ø²ÙŠÙ„ Ø§Ù„Ù„Ø§Ø­Ù‚Ø© Ø¥Ù† ÙˆÙØ¬Ø¯Øª
        return c[:-len(suf)] if c.endswith(suf) else c
    else:
        # Ù†Ø¶Ù…Ù† ÙˆØ¬ÙˆØ¯ Ø§Ù„Ù„Ø§Ø­Ù‚Ø©
        return c if c.endswith(suf) else (c + suf)


def apply_suffix_policy(raw_code: str, cfg: dict, context: str, checkbox_value: Optional[bool],
                        location: Optional[str] = None) -> str:
    """
    ğŸš€ AICR v3.0 - Auto Intelligent Code Resolver
    Ø£Ù‚ÙˆÙ‰ Ø®ÙˆØ§Ø±Ø²Ù…ÙŠØ© Ù„ØªØµÙ†ÙŠÙ Ø§Ù„ÙƒÙˆØ¯ (Ø£ØµÙ„ÙŠ / ØªØ¬Ø§Ø±ÙŠ) ØªÙ„Ù‚Ø§Ø¦ÙŠÙ‹Ø§ Ø¨Ø¯Ù‚Ø© Ø¹Ø§Ù„ÙŠØ©
    Ù…Ø¹ ÙƒØ´Ù Ø§Ù„ØªÙ†Ø§Ù‚Ø¶Ø§Øª ÙˆØªÙ†Ø¨ÙŠÙ‡ Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…ÙŠÙ†
    """

    base = _normalize_code_text(_extract_code_from_text(raw_code) or raw_code, cfg, context=context)
    suf = _suffix_to_use(cfg)
    orig_code = base
    comm_code = base + suf

    # âœ… Ù‚Ø±Ø§Ø¡Ø© Ø§Ù„Ù…Ø®Ø²ÙˆÙ†
    try:
        stock, _, _, _ = read_all()
        df = stock.copy()
    except Exception:
        st.error("âŒ Ø®Ø·Ø£ ÙÙŠ Ù‚Ø±Ø§Ø¡Ø© Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù…Ø®Ø²ÙˆÙ†.")
        return base

    # ÙÙ„ØªØ±Ø© Ø­Ø³Ø¨ Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ù„Ùˆ Ù…ÙˆØ¬ÙˆØ¯
    if location:
        df_site = df[df["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] == location]
    else:
        df_site = df

    codes = set(df_site["Ø§Ù„ÙƒÙˆØ¯"].astype(str).tolist())
    all_codes = set(df["Ø§Ù„ÙƒÙˆØ¯"].astype(str).tolist())

    has_orig_site = orig_code in codes
    has_comm_site = comm_code in codes
    has_orig_global = orig_code in all_codes
    has_comm_global = comm_code in all_codes

    # ğŸ” Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø°ÙƒØ§Ø¡
    want_original = None
    confidence = 0.0
    reason = ""

    # ğŸ§© Ù…Ø±Ø­Ù„Ø© Ø§Ù„ØªØ­Ù„ÙŠÙ„
    if checkbox_value is not None:
        want_original = bool(checkbox_value)
        confidence = 1.0
        reason = "Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… Ø­Ø¯Ù‘Ø¯ ÙŠØ¯ÙˆÙŠÙ‹Ø§ âœ…"

    elif has_orig_site and not has_comm_site:
        want_original = True
        confidence = 0.95
        reason = "Ø§Ù„ÙƒÙˆØ¯ Ø§Ù„Ø£ØµÙ„ÙŠ Ù…ÙˆØ¬ÙˆØ¯ ÙÙ‚Ø· ÙÙŠ Ù‡Ø°Ø§ Ø§Ù„Ù…ÙˆÙ‚Ø¹"

    elif not has_orig_site and has_comm_site:
        want_original = False
        confidence = 0.95
        reason = "Ø§Ù„ÙƒÙˆØ¯ Ø§Ù„ØªØ¬Ø§Ø±ÙŠ ÙÙ‚Ø· Ù…ÙˆØ¬ÙˆØ¯ ÙÙŠ Ù‡Ø°Ø§ Ø§Ù„Ù…ÙˆÙ‚Ø¹"

    elif has_orig_site and has_comm_site:
        want_original = None
        confidence = 0.0
        reason = "ÙƒÙ„Ø§ Ø§Ù„Ù†Ø³Ø®ØªÙŠÙ† Ù…ÙˆØ¬ÙˆØ¯ØªØ§Ù† Ø¨Ù†ÙØ³ Ø§Ù„Ù…ÙˆÙ‚Ø¹ âš ï¸"

    elif not has_orig_site and not has_comm_site:
        # ğŸ‘ï¸ ÙØ­Øµ Ø¹Ø§Ù… ÙÙŠ ÙƒÙ„ Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹
        if has_orig_global and not has_comm_global:
            want_original = True
            confidence = 0.85
            reason = "Ø§Ù„Ø£ØµÙ„ÙŠ Ù…ÙˆØ¬ÙˆØ¯ ÙÙŠ Ù…ÙˆØ§Ù‚Ø¹ Ø£Ø®Ø±Ù‰"
        elif not has_orig_global and has_comm_global:
            want_original = False
            confidence = 0.85
            reason = "Ø§Ù„ØªÙ‚Ù„ÙŠØ¯ Ù…ÙˆØ¬ÙˆØ¯ ÙÙŠ Ù…ÙˆØ§Ù‚Ø¹ Ø£Ø®Ø±Ù‰"
        else:
            want_original = True
            confidence = 0.6
            reason = "ÙƒÙˆØ¯ Ø¬Ø¯ÙŠØ¯ - ØªÙ… Ø§ÙØªØ±Ø§Ø¶ Ø£Ù†Ù‡ Ø£ØµÙ„ÙŠ"

            # ğŸ§  Ù…Ù†Ø·Ù‚ Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹ â€” Ø¥Ø°Ø§ Ø§Ù„Ø£ØµÙ„ÙŠ ÙˆØ§Ù„ØªØ¬Ø§Ø±ÙŠ ÙÙŠ Ù…ÙˆØ§Ù‚Ø¹ Ù…Ø®ØªÙ„ÙØ©
            orig_locs = sorted(df[df["Ø§Ù„ÙƒÙˆØ¯"] == orig_code]["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].astype(str).unique().tolist())
            comm_locs = sorted(df[df["Ø§Ù„ÙƒÙˆØ¯"] == comm_code]["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].astype(str).unique().tolist())

            if location:
                if orig_locs and comm_locs:
                    # Ø§Ù„Ù…ÙˆÙ‚Ø¹ ÙŠØ­ØªÙˆÙŠ Ø§Ù„Ø£ØµÙ„ÙŠ ÙÙ‚Ø·
                    if (location in orig_locs) and (location not in comm_locs):
                        want_original = True
                        confidence = 1.0
                        reason = "Ø§Ù„Ù…ÙˆÙ‚Ø¹ ÙŠØ­ØªÙˆÙŠ ÙÙ‚Ø· Ø§Ù„Ù†Ø³Ø®Ø© Ø§Ù„Ø£ØµÙ„ÙŠØ© â†’ Ø¥Ø¬Ø¨Ø§Ø± ØµØ­ÙŠØ­ ØªÙ„Ù‚Ø§Ø¦ÙŠ"

                    # Ø§Ù„Ù…ÙˆÙ‚Ø¹ ÙŠØ­ØªÙˆÙŠ Ø§Ù„ØªØ¬Ø§Ø±ÙŠ ÙÙ‚Ø·
                    elif (location in comm_locs) and (location not in orig_locs):
                        want_original = False
                        confidence = 1.0
                        reason = "Ø§Ù„Ù…ÙˆÙ‚Ø¹ ÙŠØ­ØªÙˆÙŠ ÙÙ‚Ø· Ø§Ù„Ù†Ø³Ø®Ø© Ø§Ù„ØªØ¬Ø§Ø±ÙŠØ© â†’ Ø¥Ø¬Ø¨Ø§Ø± ØµØ­ÙŠØ­ ØªÙ„Ù‚Ø§Ø¦ÙŠ"

    # ğŸš¨ Ù…Ø±Ø­Ù„Ø© Ø§Ù„Ø­Ù…Ø§ÙŠØ© Ù…Ù† Ø§Ù„ØªÙ†Ø§Ù‚Ø¶Ø§Øª
    duplicates = df_site[
        df_site["Ø§Ù„ÙƒÙˆØ¯"].astype(str).str.fullmatch(orig_code) | df_site["Ø§Ù„ÙƒÙˆØ¯"].astype(str).str.fullmatch(comm_code)]
    if len(duplicates) > 2:
        st.error("âš ï¸ Ø®Ø·Ø£: ØªÙ… Ø§Ù„Ø¹Ø«ÙˆØ± Ø¹Ù„Ù‰ Ø£ÙƒØ«Ø± Ù…Ù† Ø³Ø¬Ù„ Ù„Ù†ÙØ³ Ø§Ù„ÙƒÙˆØ¯ ÙÙŠ Ù‡Ø°Ø§ Ø§Ù„Ù…ÙˆÙ‚Ø¹.")
        return base

    # ğŸ¯ Ø§Ù„Ù‚Ø±Ø§Ø± Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠ
    if want_original is None:
        st.warning(f"âš ï¸ Ø§Ù„ÙƒÙˆØ¯ '{base}' Ù…ÙˆØ¬ÙˆØ¯ ÙƒØ£ØµÙ„ÙŠ ÙˆØªÙ‚Ù„ÙŠØ¯ ÙÙŠ Ù†ÙØ³ Ø§Ù„Ù…ÙˆÙ‚Ø¹. ÙŠØ±Ø¬Ù‰ ØªØ­Ø¯ÙŠØ¯ Ø§Ù„Ù†ÙˆØ¹ ÙŠØ¯ÙˆÙŠÙ‹Ø§.")
        st.dataframe(duplicates[["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„ÙƒÙ…ÙŠØ©"]])
        return base



    return ensure_original_flag(base, cfg, want_original)


# -------------------------------------------------
# ØªØ­Ù…ÙŠÙ„ Ø£ÙˆÙ„ÙŠ Ù„Ù„ÙˆØ±Ù‚Ø© (Ø¨Ø¯ÙˆÙ† Ø±Ø¤ÙˆØ³) + Ø§ÙƒØªØ´Ø§Ù Ø§Ù„Ø´Ø¨ÙƒØ©
# -------------------------------------------------
@st.cache_data(show_spinner=False)
def _load_raw_excel(path: str, _mtime: float) -> dict:
    xls = pd.ExcelFile(path, engine="openpyxl")
    sheets = {}
    for s in xls.sheet_names:
        sheets[s] = pd.read_excel(xls, sheet_name=s, header=None)
    return sheets


def _drop_all_nan(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all")
    return df


def _detect_grid(df_raw: pd.DataFrame) -> pd.DataFrame:
    df = _drop_all_nan(df_raw)
    keep = []
    for c in df.columns:
        name = str(c).strip().lower()
        if name in ["", "nan", "none", "unnamed: 0"]:
            continue
        keep.append(c)
    if keep:
        df = df[keep]
    return df.reset_index(drop=True)


# -------------------------------------------------
# ØªØ·Ø¨ÙŠØ¹ Stock / Transactions
# -------------------------------------------------
def _first_row_looks_like_header(df: pd.DataFrame) -> bool:
    try:
        s = df.iloc[0].astype(str).str.strip()
        keywords = ["ÙƒÙˆØ¯", "ÙˆØµÙ", "Ù…ÙˆÙ‚Ø¹", "Ù…Ø®Ø²ÙˆÙ†"]
        hits = sum(any(k in cell for k in keywords) for cell in s)
        return hits >= 2
    except Exception:
        return False


def _heuristic_rebuild_stock(df: pd.DataFrame) -> pd.DataFrame:
    """Ø¥Ø¹Ø§Ø¯Ø© Ø¨Ù†Ø§Ø¡ Ø§Ù„Ø¬Ø¯ÙˆÙ„ Ù…Ù† Ù†Øµ ØºÙŠØ± Ù…Ù†Ø¸Ù…."""
    df = df.copy()
    rows = []
    for _, r in df.iterrows():
        cells = [str(r[c]).strip() for c in df.columns]
        # Ø§ÙØªØ±Ø§Ø¶: Ø§Ù„Ø¹Ù…ÙˆØ¯ 0 = Ø§Ù„ÙƒÙˆØ¯ØŒ Ø§Ù„Ø¹Ù…ÙˆØ¯ 1 = Ø§Ù„ÙˆØµÙØŒ Ø§Ù„Ø¹Ù…ÙˆØ¯ 2 = Ø§Ù„Ù…ÙˆÙ‚Ø¹ØŒ Ø§Ù„Ø¹Ù…ÙˆØ¯ 3 = Ø§Ù„Ù…Ø®Ø²ÙˆÙ†
        code = cells[0] if len(cells) > 0 else ""
        desc = cells[1] if len(cells) > 1 else ""
        loc = cells[2] if len(cells) > 2 else ""
        qty_str = cells[3] if len(cells) > 3 else ""
        try:
            qty = int(float(qty_str)) if qty_str else 0
        except ValueError:
            qty = 0
        # ØªÙ†Ø¸ÙŠÙ Ø§Ù„ÙƒÙˆØ¯ Ù…Ù† Ø§Ù„Ø£Ù‚ÙˆØ§Ø³ Ø¥Ù† ÙˆÙØ¬Ø¯Øª
        if "[" in code and "]" in code:
            code_clean = code.split("[")[1].split("]")[0].strip()
            if code_clean:
                code = code_clean
        # ØªÙ†Ø¸ÙŠÙ Ø§Ù„ÙˆØµÙ
        if "[" in desc and "]" in desc:
            desc_clean = desc.split("]", 1)[1].strip()
            if desc_clean:
                desc = desc_clean
        rows.append({"Ø§Ù„ÙƒÙˆØ¯": code, "Ø§Ù„ÙˆØµÙ": desc, "Ø§Ù„Ù…ÙˆÙ‚Ø¹": loc, "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†": qty})
    out = pd.DataFrame(rows).dropna(how="all")
    out["Ø§Ù„ÙƒÙˆØ¯"] = out["Ø§Ù„ÙƒÙˆØ¯"].fillna("").astype(str).str.strip()
    out["Ø§Ù„ÙˆØµÙ"] = out["Ø§Ù„ÙˆØµÙ"].fillna("").astype(str).str.strip()
    out["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] = out["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].fillna("").astype(str).str.strip()
    out["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"] = pd.to_numeric(out["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"], errors="coerce").fillna(0).astype(int)
    mask_empty = (out[["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]].astype(str).apply(lambda s: s.str.len()) == 0).all(axis=1)
    return out[~mask_empty].reset_index(drop=True)


def _normalize_stock_cols(df: pd.DataFrame) -> pd.DataFrame:
    df0 = df.copy()
    if not df0.empty and _first_row_looks_like_header(df0):
        df0.columns = df0.iloc[0].tolist()
        df0 = df0.iloc[1:].reset_index(drop=True)
    mapping = {}
    for col in df0.columns:
        t = str(col).strip()
        if "ÙƒÙˆØ¯" in t:
            mapping[col] = "Ø§Ù„ÙƒÙˆØ¯"
        elif "ÙˆØµÙ" in t:
            mapping[col] = "Ø§Ù„ÙˆØµÙ"
        elif "Ù…ÙˆÙ‚Ø¹" in t:
            mapping[col] = "Ø§Ù„Ù…ÙˆÙ‚Ø¹"
        elif "Ù…Ø®Ø²ÙˆÙ†" in t:
            mapping[col] = "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"
    if mapping:
        df0 = df0.rename(columns=mapping)
        required = ["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ù…ÙˆÙ‚Ø¹", "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"]
        if any(c not in df0.columns for c in required):
            df0 = _heuristic_rebuild_stock(df0)
        else:
            df0 = df0.dropna(subset=["Ø§Ù„Ù…ÙˆÙ‚Ø¹"])
            df0["Ø§Ù„ÙƒÙˆØ¯"] = df0["Ø§Ù„ÙƒÙˆØ¯"].fillna("").astype(str).str.strip()
            df0["Ø§Ù„ÙˆØµÙ"] = df0["Ø§Ù„ÙˆØµÙ"].fillna("").astype(str).str.strip()
            df0["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] = df0["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].fillna("").astype(str).str.strip()
            df0["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"] = pd.to_numeric(df0["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"], errors="coerce").fillna(0).astype(int)
    else:
        df0 = _heuristic_rebuild_stock(df0)
    return df0[["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ù…ÙˆÙ‚Ø¹", "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"]].reset_index(drop=True)


def _normalize_tx_cols(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„Ù†ÙˆØ¹", "Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ù…Ù†_Ù…ÙˆÙ‚Ø¹", "Ø§Ù„Ù‰_Ù…ÙˆÙ‚Ø¹", "Ø§Ù„ÙƒÙ…ÙŠØ©", "Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…", "Ù…Ù„Ø§Ø­Ø¸Ø©"]
    if df.empty:
        return pd.DataFrame(columns=cols)
    if df.iloc[0].astype(str).str.contains("Ø§Ù„ØªØ§Ø±ÙŠØ®|Ø§Ù„Ù†ÙˆØ¹|Ø§Ù„ÙƒÙˆØ¯").any():
        df.columns = df.iloc[0].tolist()
        df = df.iloc[1:]
    mapping = {}
    for col in df.columns:
        t = str(col).strip()
        if "ØªØ§Ø±ÙŠØ®" in t:
            mapping[col] = "Ø§Ù„ØªØ§Ø±ÙŠØ®"
        elif "Ù†ÙˆØ¹" in t:
            mapping[col] = "Ø§Ù„Ù†ÙˆØ¹"
        elif "ÙƒÙˆØ¯" in t:
            mapping[col] = "Ø§Ù„ÙƒÙˆØ¯"
        elif "ÙˆØµÙ" in t:
            mapping[col] = "Ø§Ù„ÙˆØµÙ"
        elif "Ù…Ù†" in t and "Ù…ÙˆÙ‚Ø¹" in t:
            mapping[col] = "Ù…Ù†_Ù…ÙˆÙ‚Ø¹"
        elif "Ø§Ù„Ù‰" in t and "Ù…ÙˆÙ‚Ø¹" in t:
            mapping[col] = "Ø§Ù„Ù‰_Ù…ÙˆÙ‚Ø¹"
        elif "ÙƒÙ…ÙŠØ©" in t:
            mapping[col] = "Ø§Ù„ÙƒÙ…ÙŠØ©"
        elif "Ù…Ø³ØªØ®Ø¯Ù…" in t:
            mapping[col] = "Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…"
        elif "Ù…Ù„Ø§Ø­Ø¸" in t:
            mapping[col] = "Ù…Ù„Ø§Ø­Ø¸Ø©"
    df = df.rename(columns=mapping)
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df["Ø§Ù„ÙƒÙ…ÙŠØ©"] = pd.to_numeric(df["Ø§Ù„ÙƒÙ…ÙŠØ©"], errors="coerce").fillna(0).astype(int)
    return df[cols].reset_index(drop=True)


# -------------------------------------------------
# Ù‚Ø±Ø§Ø¡Ø©/ÙƒØªØ§Ø¨Ø© Ù…ÙˆØ­Ù‘Ø¯Ø© + ØªÙ„ÙˆÙŠÙ† Ù…Ù„Ù Ø§Ù„Ø¥ÙƒØ³Ù„ Ø¨Ø¹Ø¯ ÙƒÙ„ Ø­ÙØ¸
# -------------------------------------------------
def _compact_stock(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df["Ø§Ù„ÙƒÙˆØ¯"] = df["Ø§Ù„ÙƒÙˆØ¯"].fillna("").astype(str).str.strip()
    df["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] = df["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].fillna("").astype(str).str.strip()
    df["Ø§Ù„ÙˆØµÙ"] = df["Ø§Ù„ÙˆØµÙ"].fillna("").astype(str).str.strip()
    df["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"] = pd.to_numeric(df["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"], errors="coerce").fillna(0).astype(int)
    out = (df.groupby(["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"], as_index=False)
           .agg(Ø§Ù„Ù…Ø®Ø²ÙˆÙ†=("Ø§Ù„Ù…Ø®Ø²ÙˆÙ†", "sum"), Ø§Ù„ÙˆØµÙ=("Ø§Ù„ÙˆØµÙ", "first")))
    return out[["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ù…ÙˆÙ‚Ø¹", "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"]].sort_values(["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]).reset_index(drop=True)


def _apply_global_code_normalization(df: pd.DataFrame, context: str):
    cfg = load_config()
    if df.empty:
        return df
    df = df.copy()
    df["Ø§Ù„ÙƒÙˆØ¯"] = df["Ø§Ù„ÙƒÙˆØ¯"].apply(lambda s: _normalize_code_text(s, cfg, context=context))
    return df


def _header_col_index(ws, header_text: str) -> Optional[int]:
    for cell in ws[1]:
        if str(cell.value).strip() == header_text:
            return cell.column
    return None


def _apply_excel_coloring(path: str):
    try:
        cfg = load_config()
        min_level = int(cfg.get("global_min_level", 2))
        suf = _suffix_to_use(cfg)
        wb = load_workbook(path)
        if "Stock" not in wb.sheetnames:
            wb.save(path);
            return
        ws = wb["Stock"]
        c_code = _header_col_index(ws, "Ø§Ù„ÙƒÙˆØ¯")
        c_qty = _header_col_index(ws, "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†")
        if not c_code or not c_qty:
            wb.save(path);
            return
        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fill_clear = PatternFill()
        max_row = ws.max_row
        for r in range(2, max_row + 1):
            cell_code = ws.cell(row=r, column=c_code)
            cell_qty = ws.cell(row=r, column=c_qty)
            cell_code.fill = fill_clear
            cell_qty.fill = fill_clear
            code_val = str(cell_code.value or "").strip()
            if code_val:
                # âœ… Ø§Ù„Ø¢Ù†: Ø§Ù„Ø£ØµÙ„ÙŠ (Ø¨Ø¯ÙˆÙ† -S) = Ø£Ø®Ø¶Ø±ØŒ Ø§Ù„ØªÙ‚Ù„ÙŠØ¯ (Ù…Ø¹ -S) = Ø¨Ø±ØªÙ‚Ø§Ù„ÙŠ
                if not code_val.endswith(suf):
                    cell_code.fill = fill_green
                else:
                    cell_code.fill = fill_orange
            try:
                q = int(float(cell_qty.value or 0))
            except Exception:
                q = 0
            if q <= 0:
                cell_qty.fill = fill_red
            elif q <= min_level:
                cell_qty.fill = fill_yellow
        wb.save(path)
    except Exception:
        pass


def read_all(preferred_sheet: Optional[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, List[str]]:
    ensure_excel_file()
    _drop_sheet_if_exists(EXCEL_PATH, "MinLevels")
    mtime = _file_mtime(EXCEL_PATH)
    sheets_raw = _load_raw_excel(EXCEL_PATH, mtime)
    names = list(sheets_raw.keys())
    candidate = "Stock" if "Stock" in sheets_raw else (preferred_sheet or names[0])
    stock_raw = _detect_grid(sheets_raw[candidate])
    stock = _normalize_stock_cols(stock_raw)
    if "Transactions" in sheets_raw:
        tx = _normalize_tx_cols(_detect_grid(sheets_raw["Transactions"]))
    else:
        tx = pd.DataFrame(
            columns=["Ø§Ù„ØªØ§Ø±ÙŠØ®", "Ø§Ù„Ù†ÙˆØ¹", "Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ù…Ù†_Ù…ÙˆÙ‚Ø¹", "Ø§Ù„Ù‰_Ù…ÙˆÙ‚Ø¹", "Ø§Ù„ÙƒÙ…ÙŠØ©", "Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…", "Ù…Ù„Ø§Ø­Ø¸Ø©"])
    minlvl = pd.DataFrame(columns=["Ø§Ù„ÙƒÙˆØ¯", "Ø­Ø¯_Ø¥Ø¹Ø§Ø¯Ø©_Ø§Ù„Ø·Ù„Ø¨"])
    stock = _apply_global_code_normalization(stock, context="import")
    stock = _compact_stock(stock)
    return stock, minlvl, tx, names


def write_all(stock: pd.DataFrame, _minlvl_unused: pd.DataFrame, tx: pd.DataFrame):
    stock = _compact_stock(stock)
    _backup_if_needed()

    def _write(path):
        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as w:
            stock.to_excel(w, index=False, sheet_name="Stock")
            tx.to_excel(w, index=False, sheet_name="Transactions")

    with SimpleFileLock(EXCEL_PATH):
        _atomic_write_excel(_write, EXCEL_PATH)
        _apply_excel_coloring(EXCEL_PATH)


# -------------------------------------------------
# Ø¯ÙˆØ§Ù„ Ù…Ø¬Ø§Ù„ Ø§Ù„Ø¹Ù…Ù„
# -------------------------------------------------
def get_unique_locations(stock: pd.DataFrame) -> List[str]:
    return sorted(stock["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].astype(str).unique().tolist())


def get_unique_codes(stock: pd.DataFrame) -> List[str]:
    return sorted(stock["Ø§Ù„ÙƒÙˆØ¯"].astype(str).unique().tolist())


def get_part_desc(stock: pd.DataFrame, code: str) -> str:
    m = stock[stock["Ø§Ù„ÙƒÙˆØ¯"] == code]
    return "" if m.empty else str(m["Ø§Ù„ÙˆØµÙ"].iloc[0])


def get_qty(stock: pd.DataFrame, code: str, location: str) -> int:
    m = stock[(stock["Ø§Ù„ÙƒÙˆØ¯"] == code) & (stock["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] == location)]
    return 0 if m.empty else int(m["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"].iloc[0])


def get_locations_for_code(stock: pd.DataFrame, code: str) -> List[str]:
    return sorted(stock[stock["Ø§Ù„ÙƒÙˆØ¯"] == code]["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].unique().tolist())


def set_qty(stock: pd.DataFrame, code: str, location: str, qty: int) -> pd.DataFrame:
    cfg = load_config()
    code = _normalize_code_text(code, cfg, context="ops")
    location = ("" if location is None else str(location)).strip()
    mask = (stock["Ø§Ù„ÙƒÙˆØ¯"] == code) & (stock["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] == location)
    if mask.any():
        stock.loc[mask, "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"] = int(qty)
    else:
        desc = get_part_desc(stock, code)
        new_row = {"Ø§Ù„ÙƒÙˆØ¯": code, "Ø§Ù„ÙˆØµÙ": desc, "Ø§Ù„Ù…ÙˆÙ‚Ø¹": location, "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†": int(qty)}
        stock = pd.concat([stock, pd.DataFrame([new_row])], ignore_index=True)
    return stock


def add_qty(stock: pd.DataFrame, code: str, location: str, delta: int) -> Tuple[pd.DataFrame, int]:
    current = get_qty(stock, code, location)
    new_qty = current + delta
    if new_qty < 0:
        raise ValueError("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø£Ù† ØªØµØ¨Ø­ Ø§Ù„ÙƒÙ…ÙŠØ© Ø³Ø§Ù„Ø¨Ø©.")
    stock = set_qty(stock, code, location, new_qty)
    return stock, new_qty


def append_txn(tx: pd.DataFrame, t_type: str, code: str, desc: str, qty: int,
               from_loc: Optional[str], to_loc: Optional[str],
               user: str = "", note: str = "") -> pd.DataFrame:
    new_row = {
        "Ø§Ù„ØªØ§Ø±ÙŠØ®": now_iso(),
        "Ø§Ù„Ù†ÙˆØ¹": t_type,
        "Ø§Ù„ÙƒÙˆØ¯": code,
        "Ø§Ù„ÙˆØµÙ": desc,
        "Ù…Ù†_Ù…ÙˆÙ‚Ø¹": from_loc,
        "Ø§Ù„Ù‰_Ù…ÙˆÙ‚Ø¹": to_loc,
        "Ø§Ù„ÙƒÙ…ÙŠØ©": int(qty),
        "Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…": user,
        "Ù…Ù„Ø§Ø­Ø¸Ø©": note,
    }
    return pd.concat([tx, pd.DataFrame([new_row])], ignore_index=True)


# -------------------------------------------------
# ØªÙ†Ø¨ÙŠÙ‡Ø§Øª
# -------------------------------------------------
def compute_low_and_oos(stock: pd.DataFrame, min_level: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if stock.empty:
        return (pd.DataFrame(columns=["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ"]),
                pd.DataFrame(columns=["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ"]))
    agg = stock.groupby("Ø§Ù„ÙƒÙˆØ¯", as_index=False).agg(Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ=("Ø§Ù„Ù…Ø®Ø²ÙˆÙ†", "sum"),
                                                     Ø§Ù„ÙˆØµÙ=("Ø§Ù„ÙˆØµÙ", "first"))
    oos_df = agg[agg["Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ"] <= 0].sort_values("Ø§Ù„ÙƒÙˆØ¯")
    low_df = agg[(agg["Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ"] > 0) & (agg["Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ"] <= int(min_level))].sort_values("Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ")
    return low_df[["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ"]], oos_df[["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ"]]


# -------------------------------------------------
# Ø¨Ø­Ø« Ø¨Ø³ÙŠØ· (Ù„Ù„Ø§Ø³ØªØ®Ø¯Ø§Ù… ÙÙŠ ØµÙØ­Ø© Ø§Ù„Ø¨Ø­Ø«) â€” âœ… Ù…Ø¹ Ø²Ø± "Ø£ØµÙ„ÙŠØŸ"
# -------------------------------------------------
def _parse_locations_text(loc_text: str) -> List[str]:
    tokens = [t.strip() for t in re.split(r"[,\n]+", (loc_text or "")) if t.strip()]
    return _unique_order(tokens)


def _apply_search(stock: pd.DataFrame, query_code: str, selected_locs: List[str], cfg: dict,
                  exact_code: bool = True, is_orig: bool = True) -> pd.DataFrame:
    df = stock.copy()
    if selected_locs:
        df = df[df["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].isin(selected_locs)]
    q = (_to_ascii_digits(query_code or "")).strip()
    if q:
        # âœ… ØªØ·Ø¨ÙŠÙ‚ Ø³ÙŠØ§Ø³Ø© Ø§Ù„Ù„Ø§Ø­Ù‚Ø© Ø¹Ù„Ù‰ Ø§Ù„ÙƒÙˆØ¯ Ø§Ù„Ù…Ø¯Ø®Ù„ ÙÙŠ Ø§Ù„Ø¨Ø­Ø«
        norm_q = apply_suffix_policy(q, cfg, context="scan", checkbox_value=is_orig)
        if exact_code:
            df = df[df["Ø§Ù„ÙƒÙˆØ¯"].astype(str).str.strip() == norm_q.strip()]
        else:
            df = df[
                df["Ø§Ù„ÙƒÙˆØ¯"].astype(str).str.contains(norm_q, case=False, na=False) |
                df["Ø§Ù„ÙˆØµÙ"].astype(str).str.contains(norm_q, case=False, na=False)
                ]
    elif selected_locs:
        # Ø¹Ø±Ø¶ ÙƒÙ„ Ø§Ù„Ù‚Ø·Ø¹ ÙÙŠ Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ù…Ø­Ø¯Ø¯ (Ø¨Ø¯ÙˆÙ† ÙƒÙˆØ¯)
        pass
    else:
        df = pd.DataFrame(columns=["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ù…ÙˆÙ‚Ø¹", "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"])
    return df.reset_index(drop=True)


def _summary_by_code(df: pd.DataFrame, min_level: int) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ", "Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹", "Ø§Ù„Ø­Ø§Ù„Ø©"])
    out = df.groupby("Ø§Ù„ÙƒÙˆØ¯", as_index=False).agg(
        Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ=("Ø§Ù„Ù…Ø®Ø²ÙˆÙ†", "sum"),
        Ø§Ù„ÙˆØµÙ=("Ø§Ù„ÙˆØµÙ", "first"),
        Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹=("Ø§Ù„Ù…ÙˆÙ‚Ø¹", lambda x: ", ".join(sorted(x.astype(str).unique())))
    )

    def status(q):
        if q <= 0: return "ØºÙŠØ± Ù…ØªÙˆÙØ±"
        if q <= min_level: return "Ù…Ù†Ø®ÙØ¶"
        return "Ø¬ÙŠØ¯"

    out["Ø§Ù„Ø­Ø§Ù„Ø©"] = out["Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ"].apply(status)
    return out.sort_values(["Ø§Ù„Ø­Ø§Ù„Ø©", "Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ", "Ø§Ù„ÙƒÙˆØ¯"]).reset_index(drop=True)


def _lookup_code(stock: pd.DataFrame, code: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = stock[stock["Ø§Ù„ÙƒÙˆØ¯"] == code].copy()
    if df.empty:
        return df, pd.DataFrame(columns=["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„ÙˆØµÙ", "Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ"])
    s = df.groupby("Ø§Ù„ÙƒÙˆØ¯", as_index=False).agg(Ø§Ù„Ø¥Ø¬Ù…Ø§Ù„ÙŠ=("Ø§Ù„Ù…Ø®Ø²ÙˆÙ†", "sum"), Ø§Ù„ÙˆØµÙ=("Ø§Ù„ÙˆØµÙ", "first"))
    return df.sort_values("Ø§Ù„Ù…ÙˆÙ‚Ø¹"), s


# -------------------------------------------------
# ØªÙ†Ù‚Ù‘Ù„ Ø¨ÙŠÙ† Ø§Ù„ØµÙØ­Ø§Øª
# -------------------------------------------------
PAGES = [
    "Ù„ÙˆØ­Ø© Ø§Ù„ØªØ­ÙƒÙ…", "Ø¨Ø­Ø«/Ù…Ø³Ø­", "Ø§Ù„Ø¹Ù…Ù„ÙŠØ§Øª", "Ø§Ù„Ø¬Ø±Ø¯",
    "Ø¥Ø¶Ø§ÙØ© Ù‚Ø·Ø¹Ø© Ø¬Ø¯ÙŠØ¯Ø©",
    "Ø¯Ù…Ø¬ Ù…Ù„Ù Ø¬Ø¯ÙŠØ¯", "ØªØ­Ø±ÙŠØ± Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª", "Ø§Ø³ØªÙŠØ±Ø§Ø¯/ØªØµØ¯ÙŠØ±", "Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª"
]


def nav_to(page_name: str):
    st.session_state.menu = page_name
    st.rerun()

# -------------------------------------------------
# Ø£Ø¯ÙˆØ§Øª Ø­Ø§Ù„Ø© Ø§Ù„Ù…Ù„Ù
# -------------------------------------------------
def file_status_badge():
    try:
        ok = os.path.exists(EXCEL_PATH)
        size = os.path.getsize(EXCEL_PATH) if ok else 0
        mtime = datetime.fromtimestamp(os.path.getmtime(EXCEL_PATH)).strftime("%Y-%m-%d %H:%M:%S") if ok else "-"
        st.caption(f"ğŸ“„ Ù…Ù„Ù Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª: {'Ù…ÙˆØ¬ÙˆØ¯' if ok else 'ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯'} | Ø§Ù„Ø­Ø¬Ù…: {size:,} Ø¨Ø§ÙŠØª | Ø¢Ø®Ø± ØªØ¹Ø¯ÙŠÙ„: {mtime}")
    except Exception as e:
        st.caption(f"âš ï¸ ØªØ¹Ø°Ù‘Ø± Ù‚Ø±Ø§Ø¡Ø© Ø­Ø§Ù„Ø© Ø§Ù„Ù…Ù„Ù: {e}")


# -------------------------------------------------
# ØµÙØ­Ø©: Ø¨Ø­Ø«/Ù…Ø³Ø­ â€” âœ… Ù…Ø¹ Ø²Ø± "Ø£ØµÙ„ÙŠØŸ"
# -------------------------------------------------
def page_find_and_scan():
    st.subheader("Ø¨Ø­Ø« Ø¯Ø§Ø®Ù„ Ø§Ù„Ù…Ø®Ø²Ù† (Ø¨Ø¯ÙˆÙ† Ø§Ø³ØªÙ„Ø§Ù…)")

    file_status_badge()
    stock, minlvl, tx, _ = read_all()
    cfg = load_config()
    min_level = int(cfg.get("global_min_level", 2))

    loc_text = st.text_input("ÙÙ„ØªØ±Ø© Ø¨Ø§Ù„Ù…ÙˆÙ‚Ø¹ (ÙŠÙ…ÙƒÙ† Ø¥Ø¯Ø®Ø§Ù„ Ø¹Ø¯Ø© Ù…ÙˆØ§Ù‚Ø¹ Ø¨ÙÙˆØ§ØµÙ„ Ø£Ùˆ Ø£Ø³Ø·Ø±)", value="", key="simple_loc_text",
                             placeholder="Ù…Ø«Ø§Ù„: Ø±Ù-Ø£1, ØµÙ†Ø¯ÙˆÙ‚-2")
    selected_locs = _parse_locations_text(loc_text)

    col_a, col_b, col_c = st.columns([3, 3, 1])
    with col_a:
        manual_code = st.text_input("Ø§Ù„ÙƒÙˆØ¯ (ÙƒØªØ§Ø¨ÙŠ)", key="manual_code_input",
                                    placeholder="Ø§ÙƒØªØ¨ Ø§Ù„ÙƒÙˆØ¯ Ø£Ùˆ Ø§ØªØ±ÙƒÙ‡ ÙØ§Ø±ØºÙ‹Ø§ Ù„Ø¹Ø±Ø¶ ÙƒÙ„ Ø§Ù„Ù‚Ø·Ø¹ ÙÙŠ Ø§Ù„Ù…ÙˆÙ‚Ø¹")
    with col_b:
        st.caption("Ø¶Ø¹ Ø§Ù„Ù…Ø¤Ø´Ø± Ù‡Ù†Ø§ Ø«Ù… Ø§Ù…Ø³Ø­ Ø§Ù„Ø¨Ø§Ø±ÙƒÙˆØ¯ Ø£Ùˆ Ø£Ø¯Ø®Ù„ Ø§Ù„ÙƒÙˆØ¯ ÙˆØ§Ø¶ØºØ· Enter.")

        def _on_scan():
            raw = st.session_state.get("scanner_code_input", "")
            st.session_state.scanner_code_input = ""
            st.session_state.last_search_code = raw

        st.text_input("Ø§Ù„ÙƒÙˆØ¯ (Ù…Ø§Ø³Ø­ Ø¶ÙˆØ¦ÙŠ)", key="scanner_code_input", on_change=_on_scan)
    with col_c:
        st.markdown('<div class="orig-checkbox">', unsafe_allow_html=True)
        is_orig = st.checkbox("Ø£ØµÙ„ÙŠØŸ", value=True, key="search_orig")
        st.markdown('</div>', unsafe_allow_html=True)

    search_code = st.session_state.get("last_search_code", "").strip() or manual_code.strip()
    filtered = _apply_search(stock, search_code, selected_locs, cfg=cfg, exact_code=bool(search_code), is_orig=is_orig)

    if selected_locs and not search_code:
        filtered = stock[stock["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].isin(selected_locs)].copy().reset_index(drop=True)

    st.markdown("**Ø§Ù„Ù…Ù„Ø®Øµ Ø­Ø³Ø¨ Ø§Ù„ÙƒÙˆØ¯:**")
    st.dataframe(_summary_by_code(filtered, min_level), use_container_width=True, height=180)
    st.markdown("**ØªÙØ§ØµÙŠÙ„ Ø­Ø³Ø¨ Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹:**")
    st.dataframe(filtered.sort_values(["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]), use_container_width=True, height=320)
    if not filtered.empty and not search_code and selected_locs:
        st.info(f"Ø¹Ø±Ø¶ Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù‚Ø·Ø¹ ÙÙŠ Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹: {', '.join(selected_locs)}")
    elif search_code and filtered.empty:
        st.info("Ù„Ø§ ØªÙˆØ¬Ø¯ Ù†ØªØ§Ø¦Ø¬ Ù„Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯ Ø¶Ù…Ù† Ù†Ø·Ø§Ù‚ Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹ Ø§Ù„Ù…Ø­Ø¯Ø¯.")


# -------------------------------------------------
# ØµÙØ­Ø©: Ø§Ù„Ø¬Ø±Ø¯ (Ù…ÙØ­Ø³Ù‘Ù†Ø© Ø¬Ø¯Ù‹Ø§ Ù…Ø¹ Ø§Ù„ØªØ­Ù‚Ù‚ Ø§Ù„ØµØ­ÙŠØ­ Ù…Ù† Ø§Ù„Ù…ÙˆÙ‚Ø¹)
# -------------------------------------------------
def _init_stocktake_state():
    if "stocktake" not in st.session_state:
        # ğŸŸ¦ ØªØ®Ø²ÙŠÙ† Ù†ØªØ§Ø¦Ø¬ ÙƒÙ„ Ù…ÙˆÙ‚Ø¹
        if "stocktake_sites" not in st.session_state:
            st.session_state.stocktake_sites = {}  # {Ø§Ù„Ù…ÙˆÙ‚Ø¹: DataFrame}

        st.session_state.stocktake = {
            "scope": "all",
            "loc": "",
            "is_orig": True,
            "items": {},
            "manual_rev": 0,
            "scan_rev": 0,
            "last_code": "",
        }


def _scan_callback(scan_key: str):
    raw = st.session_state.get(scan_key, "")
    st.session_state.stocktake["last_code"] = raw
    st.session_state.stocktake["scan_rev"] += 1
    st.rerun()


def _clear_inputs_and_rerun():
    st.session_state.stocktake["last_code"] = ""
    st.session_state.stocktake["manual_rev"] += 1
    st.session_state.stocktake["scan_rev"] += 1
    st.rerun()


def page_stocktake():
    st.subheader("Ø§Ù„Ø¬Ø±Ø¯ Ø§Ù„Ù…Ø¨Ø³Ù‘Ø·")

    file_status_badge()
    _init_stocktake_state()
    cfg = load_config()
    stock, minlvl, tx, _ = read_all()
    min_level = int(cfg.get("global_min_level", 2))

    # Ø¥Ø¹Ø¯Ø§Ø¯ Ø§Ù„Ù†Ø·Ø§Ù‚ ÙˆØ§Ù„Ù…ÙˆÙ‚Ø¹
    c1, c2 = st.columns([2, 2])
    with c1:
        scope = st.radio("Ù†Ø·Ø§Ù‚ Ø§Ù„Ø¬Ø±Ø¯", ["Ø§Ù„Ù…Ø®Ø²Ù† ÙƒØ§Ù…Ù„", "Ø­Ø³Ø¨ Ù…ÙˆÙ‚Ø¹ Ù…Ø­Ø¯Ø¯"], horizontal=True,
                         index=0 if st.session_state.stocktake["scope"] == "all" else 1)
        st.session_state.stocktake["scope"] = "all" if scope == "Ø§Ù„Ù…Ø®Ø²Ù† ÙƒØ§Ù…Ù„" else "loc"
    with c2:
        if st.session_state.stocktake["scope"] == "loc":
            loc_input = st.text_input("Ø§Ù„Ù…ÙˆÙ‚Ø¹ (ÙƒØªØ§Ø¨ÙŠ)", value=st.session_state.stocktake.get("loc", ""),
                                      placeholder="Ù…Ø«Ø§Ù„: Ø±Ù-Ø£1", key="stk_loc_input")
            st.session_state.stocktake["loc"] = loc_input.strip()
            # ğŸŸ¦ Ù…Ù†Ø·Ù‚ Ø§Ù„Ø§Ù†ØªÙ‚Ø§Ù„ Ø¨ÙŠÙ† Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹
            prev_loc = st.session_state.stocktake.get("prev_loc", "")
            current_loc = st.session_state.stocktake["loc"].strip()

            # Ø¥Ø°Ø§ ÙƒØ§Ù† Ù‡Ù†Ø§Ùƒ Ù…ÙˆÙ‚Ø¹ Ø³Ø§Ø¨Ù‚ ÙˆØªÙ… ØªØºÙŠÙŠØ± Ø§Ù„Ù…ÙˆÙ‚Ø¹
            if prev_loc and current_loc and current_loc != prev_loc:
                # Ù‡Ù„ ÙŠÙˆØ¬Ø¯ Ù‚Ø·Ø¹ Ù„Ù… ØªÙØ¬Ø±Ù‘Ø¯ØŸ
                remaining = [
                    k for k in st.session_state.stocktake["items"].keys()
                    if k[1] == prev_loc
                ]

                if remaining:
                    st.toast(
                        f"âš ï¸ ÙŠÙˆØ¬Ø¯ Ù‚Ø·Ø¹ Ù„Ù… ØªÙØ¬Ø±Ù‘Ø¯ ÙÙŠ Ø§Ù„Ù…ÙˆÙ‚Ø¹ '{prev_loc}'.",
                        icon="âš ï¸",
                        duration=15
                    )
                    # Ø¥Ø¹Ø§Ø¯Ø© Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ù‚Ø¯ÙŠÙ…
                    st.session_state.stocktake["loc"] = prev_loc
                    st.rerun()
                else:
                    # Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ù…ÙƒØªÙ…Ù„ â†’ Ù†Ø­ÙØ¸Ù‡
                    df_site = pd.DataFrame([
                        {
                            "Ø§Ù„ÙƒÙˆØ¯": code,
                            "Ø§Ù„Ù…ÙˆÙ‚Ø¹": prev_loc,
                            "Ø§Ù„Ø¹Ø¯Ø¯ Ø§Ù„ÙØ¹Ù„ÙŠ": d["count"],
                            "Ø¹Ø¯Ø¯ Ø§Ù„Ù†Ø¸Ø§Ù…": d["sys_qty"],
                        }
                        for (code, loc), d in st.session_state.stocktake["items"].items()
                        if loc == prev_loc
                    ])
                    if not df_site.empty:
                        st.session_state.stocktake_sites[prev_loc] = df_site.copy()

                    # Ø¥Ø²Ø§Ù„Ø© Ø¹Ù†Ø§ØµØ± Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ø³Ø§Ø¨Ù‚ Ù…Ù† Ø§Ù„Ø³Ù„Ø©
                    st.session_state.stocktake["items"] = {
                        k: v for k, v in st.session_state.stocktake["items"].items()
                        if k[1] != prev_loc
                    }

                    st.toast(
                        f"âœ… ØªÙ… Ø­ÙØ¸ Ø¬Ø±Ø¯ Ø§Ù„Ù…ÙˆÙ‚Ø¹ '{prev_loc}'. ÙŠÙ…ÙƒÙ†Ùƒ Ø§Ù„Ø¨Ø¯Ø¡ ÙÙŠ '{current_loc}'.",
                        icon="âœ…",
                        duration=8
                    )

            # ØªØ­Ø¯ÙŠØ« Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ø³Ø§Ø¨Ù‚ Ø¯Ø§Ø¦Ù…Ø§Ù‹
            st.session_state.stocktake["prev_loc"] = current_loc

        else:
            st.text_input("Ø§Ù„Ù…ÙˆÙ‚Ø¹ (Ù…Ø¹Ø·Ù‘Ù„ ÙÙŠ ÙˆØ¶Ø¹ Ø§Ù„Ù…Ø®Ø²Ù† ÙƒØ§Ù…Ù„)", value="", disabled=True)

    # Ù…Ø¯Ø®Ù„Ø§Øª Ø§Ù„ÙƒÙˆØ¯ ÙˆØ§Ù„Ø¹Ø¯Ø¯
    c3, c4, c5 = st.columns([1, 3, 3])
    with c3:
        is_orig = st.checkbox("Ø£ØµÙ„ÙŠØŸ", value=st.session_state.stocktake.get("is_orig", True), key="stocktake_orig")
        st.session_state.stocktake["is_orig"] = is_orig
    with c4:
        manual_key = f"stk_manual_code_{st.session_state.stocktake['manual_rev']}"
        manual_code = st.text_input("Ø§Ù„ÙƒÙˆØ¯ (ÙƒØªØ§Ø¨ÙŠ)", key=manual_key, placeholder="Ø§ÙƒØªØ¨ Ø§Ù„ÙƒÙˆØ¯ Ø£Ùˆ Ø§Ù…Ø³Ø­ Ø§Ù„Ø¨Ø§Ø±ÙƒÙˆØ¯")
    with c5:
        scan_key = f"stk_scanner_code_{st.session_state.stocktake['scan_rev']}"
        st.text_input("Ø§Ù„ÙƒÙˆØ¯ (Ù…Ø§Ø³Ø­ Ø¶ÙˆØ¦ÙŠ)", key=scan_key,
                      on_change=_scan_callback, args=(scan_key,),
                      placeholder="Ø§Ù…Ø³Ø­ Ø§Ù„Ø¨Ø§Ø±ÙƒÙˆØ¯ Ù‡Ù†Ø§ Ø«Ù… Enter")

    qty = st.number_input("Ø§Ù„Ø¹Ø¯Ø¯ Ø§Ù„ÙØ¹Ù„ÙŠ", min_value=0, value=0, step=1, key="stk_count_simple")

    # ================================
    #  ğŸŸ¦ ÙˆØ¶Ø¹ Ø§Ù„Ù…Ø³Ø­ Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠ Auto Scan
    # ================================
    def _auto_scan_handler():
        raw = st.session_state.get("autoscan_input", "").strip()
        st.session_state.autoscan_input = ""

        if not raw:
            return

        st.session_state.stocktake["last_code"] = raw
        st.session_state.run_add_to_basket = True

    st.text_input(
        "ğŸŸ¦ Ø§Ù„Ù…Ø³Ø­ Ø§Ù„ØªÙ„Ù‚Ø§Ø¦ÙŠ (Ù…Ø§Ø³Ø­ Ø¶ÙˆØ¦ÙŠ Ø¨Ø¯ÙˆÙ† Ø²Ø±)",
        key="autoscan_input",
        placeholder="Ø§Ù…Ø³Ø­ Ø§Ù„Ø¨Ø§Ø±ÙƒÙˆØ¯ ÙˆØ³ÙŠØ¶Ø§Ù ØªÙ„Ù‚Ø§Ø¦ÙŠÙ‹Ø§...",
        on_change=_auto_scan_handler,
    )

    # Ø²Ø± Ø§Ù„Ø¥Ø¶Ø§ÙØ©
    pressed = st.button("Ø¥Ø¶Ø§ÙØ© Ø¥Ù„Ù‰ Ø³Ù„Ø© Ø§Ù„Ø¬Ø±Ø¯")

    if st.session_state.get("run_add_to_basket"):
        pressed = True
        st.session_state.run_add_to_basket = False

    if pressed:

        raw = st.session_state.stocktake["last_code"]

        # âœ” ØªØ·Ø¨ÙŠÙ‚ Ù…Ù†Ø·Ù‚ Ø§Ù„Ø£ØµÙ„ÙŠ/Ø§Ù„ØªØ¬Ø§Ø±ÙŠ
        code_with_suffix = apply_suffix_policy(
            raw,
            cfg,
            context="stocktake",
            checkbox_value=st.session_state.stocktake["is_orig"]
        )
        code_normalized = _normalize_code_text(code_with_suffix, cfg, context="stocktake")

        suf = _suffix_to_use(cfg)
        candidates = {code_normalized}
        if code_normalized.endswith(suf):
            candidates.add(code_normalized[:-len(suf)])
        else:
            candidates.add(code_normalized + suf)

        # --- ğŸ” Ù…Ù†Ø·Ù‚ ØªÙ…ÙŠÙŠØ² Ø§Ù„Ø£ØµÙ„ÙŠ / Ø§Ù„ØªØ¬Ø§Ø±ÙŠ Ø¨Ø¯Ù‚Ø© 100% ---
        code_norm = _normalize_code_text(raw, cfg, context="stocktake")
        suf = _suffix_to_use(cfg)

        if st.session_state.stocktake["is_orig"]:
            # Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… Ø§Ø®ØªØ§Ø± Ø£ØµÙ„ÙŠ â†’ Ø§Ù„ÙƒÙˆØ¯ ÙŠØ¬Ø¨ Ø£Ù† ÙŠÙƒÙˆÙ† Ø¨Ø¯ÙˆÙ† S
            final_code = code_norm if not code_norm.endswith(suf) else code_norm[:-len(suf)]
        else:
            # Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… Ø§Ø®ØªØ§Ø± ØªØ¬Ø§Ø±ÙŠ â†’ ÙŠØ¶ÙŠÙ S ØªÙ„Ù‚Ø§Ø¦ÙŠÙ‹Ø§ Ø¥Ø°Ø§ Ù…Ø§ ÙƒØ§Ù†Øª Ù…ÙˆØ¬ÙˆØ¯Ø©
            final_code = code_norm if code_norm.endswith(suf) else code_norm + suf

        # Ø§Ù„Ø¨Ø­Ø« Ø§Ù„Ø¯Ù‚ÙŠÙ‚ Ø¯Ø§Ø®Ù„ Ù…Ù„Ù Ø§Ù„Ø¥ÙƒØ³Ù„ â€” Ø¨Ø¯ÙˆÙ† Ø¯Ù…Ø¬ Ø§Ù„Ø£ØµÙ„ÙŠ ÙˆØ§Ù„ØªØ¬Ø§Ø±ÙŠ
        matched = stock[stock["Ø§Ù„ÙƒÙˆØ¯"] == final_code]

        # Ø§Ø³ØªØ®Ø±Ø§Ø¬ Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹ Ø§Ù„Ù…Ø±ØªØ¨Ø·Ø© Ø¨Ù†ÙØ³ Ø§Ù„ÙƒÙˆØ¯ ÙÙ‚Ø·
        sys_locs = sorted(matched["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].astype(str).unique().tolist())

        sys_locs = sorted(matched["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].unique().tolist())
        loc_entered = st.session_state.stocktake["loc"] if st.session_state.stocktake["scope"] == "loc" else None

        # ØªØ­Ù‚Ù‚ Ø§Ù„Ù…ÙˆÙ‚Ø¹
        if st.session_state.stocktake["scope"] == "loc":
            if not loc_entered:
                st.toast("âŒ Ø£Ø¯Ø®Ù„ Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ø£ÙˆÙ„Ù‹Ø§.", icon="âŒ", duration=4)
                return
            if sys_locs and loc_entered not in sys_locs:
                st.toast(f"âš ï¸ Ø§Ù„Ù…ÙˆÙ‚Ø¹ '{loc_entered}' ØºÙŠØ± Ù…Ø³Ø¬Ù„ Ù„Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯.", icon="âš ï¸", duration=4)
                return

        # ÙƒÙ…ÙŠØ© Ø§Ù„Ù†Ø¸Ø§Ù…
        sys_qty = 0
        if not matched.empty:
            if loc_entered:
                sys_qty = int(matched[matched["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] == loc_entered]["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"].sum())
            else:
                sys_qty = int(matched["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"].sum())

        # ÙƒÙ…ÙŠØ© Ø§Ù„Ø¥Ø¯Ø®Ø§Ù„
        add_qty_value = int(qty)
        if add_qty_value <= 0:
            add_qty_value = 1

        items = st.session_state.stocktake["items"]
        key = (code_normalized, loc_entered if st.session_state.stocktake["scope"] == "loc" else None)

        if key in items:
            items[key]["count"] += add_qty_value
            total_for_key = items[key]["count"]
        else:
            items[key] = {"count": add_qty_value, "sys_qty": int(sys_qty)}
            total_for_key = add_qty_value

        loc_label = "Ø§Ù„Ù…Ø®Ø²Ù† ÙƒØ§Ù…Ù„" if key[1] is None else key[1]
        # Ø­Ø³Ø§Ø¨ Ø±Ù‚Ù… Ø§Ù„ØµÙ
        row_num = list(st.session_state.stocktake["items"].keys()).index(key) + 1

        # ØªÙ†Ø¨ÙŠÙ‡ ÙƒØ§Ù…Ù„ Ù„Ù…Ø¯Ø© 10 Ø«ÙˆØ§Ù†ÙŠ ÙˆÙŠØ´Ù…Ù„ Ø±Ù‚Ù… Ø§Ù„ØµÙ
        st.toast(
            f"ğŸ”„ ØªÙ… ØªØ¹Ø¯ÙŠÙ„ Ø§Ù„ØµÙ Ø±Ù‚Ù… {row_num} â€” Ø§Ù„ÙƒÙˆØ¯: {code_normalized}, Ø§Ù„Ù…ÙˆÙ‚Ø¹: {loc_label}, Ø§Ù„Ø¹Ø¯Ø¯ Ø§Ù„Ø¬Ø¯ÙŠØ¯: {total_for_key}",
            icon="ğŸ””",
            duration=10
        )

        st.markdown(
            "<script>setTimeout(()=>document.querySelectorAll('input[placeholder=\"Ø§Ù…Ø³Ø­ Ø§Ù„Ø¨Ø§Ø±ÙƒÙˆØ¯ Ù‡Ù†Ø§ Ø«Ù… Enter\"]')[0]?.focus(),300);</script>",
            unsafe_allow_html=True,
        )

    # ---------------------------------------------------------
    #  Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ù‚Ø·Ø¹ Ø­Ø³Ø¨ Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ù‚Ø¨Ù„ Ø³Ù„Ø© Ø§Ù„Ø¬Ø±Ø¯
    # ---------------------------------------------------------
    st.markdown("### Ø§Ù„Ù‚Ø·Ø¹ Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© ÙÙŠ Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ù…Ø­Ø¯Ø¯")

    if st.session_state.stocktake["scope"] == "loc":
        loc_entered = st.session_state.stocktake["loc"].strip()
        if loc_entered:
            df_loc = stock[stock["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] == loc_entered].copy()
        else:
            df_loc = stock.copy()
    else:
        df_loc = stock.copy()

    # Ø¥Ø²Ø§Ù„Ø© Ø£ÙŠ Ù‚Ø·Ø¹Ø© Ù…ÙˆØ¬ÙˆØ¯Ø© Ù…Ø³Ø¨Ù‚Ø§Ù‹ ÙÙŠ Ø³Ù„Ø© Ø§Ù„Ø¬Ø±Ø¯ (Ø§Ù†ØªÙ‚Ø§Ù„ Ø¥Ù„Ù‰ Ø§Ù„Ø³Ù„Ø©)
    items_keys = st.session_state.stocktake["items"].keys()
    codes_in_basket = [k[0] for k in items_keys]
    locs_in_basket = [k[1] for k in items_keys]

    # Ø¥Ø°Ø§ Ø³Ù„Ø© Ø§Ù„Ø¬Ø±Ø¯ Ø¨Ù†Ø·Ø§Ù‚ Ù…ÙˆÙ‚Ø¹ Ù…Ø¹ÙŠÙ† â†’ Ø§Ø³ØªØ¨Ø¹Ø§Ø¯ Ø§Ù„Ù‚Ø·Ø¹ Ø§Ù„Ù…Ù†Ù‚ÙˆÙ„Ø© ÙÙ‚Ø·
    if st.session_state.stocktake["scope"] == "loc":
        df_loc = df_loc[~(
                (df_loc["Ø§Ù„ÙƒÙˆØ¯"].astype(str).isin(codes_in_basket)) &
                (df_loc["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].astype(str).isin(locs_in_basket))
        )]
    else:
        # Ø§Ù„Ù…Ø®Ø²Ù† ÙƒØ§Ù…Ù„ â†’ Ù„Ø§ Ù†Ø¹Ø±Ø¶ Ø£ÙŠ Ù‚Ø·Ø¹Ø© Ù…ÙˆØ¬ÙˆØ¯Ø© ÙÙŠ Ø§Ù„Ø³Ù„Ø© (Ø£ÙŠ Ù…ÙˆÙ‚Ø¹)
        df_loc = df_loc[~df_loc["Ø§Ù„ÙƒÙˆØ¯"].astype(str).isin(codes_in_basket)]

    st.dataframe(
        df_loc.sort_values(["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]),
        use_container_width=True,
        height=300
    )

    # Ø¹Ø±Ø¶ Ø§Ù„Ø³Ù„Ø©
    st.markdown("### Ø³Ù„Ø© Ø§Ù„Ø¬Ø±Ø¯")
    rows = []
    for (code, loc), d in st.session_state.stocktake["items"].items():
        rows.append({
            "Ø§Ù„ÙƒÙˆØ¯": code,
            "Ø§Ù„Ù†ÙˆØ¹": "Ø£ØµÙ„ÙŠ" if is_original_code(code, cfg) else "ØªØ¬Ø§Ø±ÙŠ",
            "Ø§Ù„Ù…ÙˆÙ‚Ø¹": "Ø§Ù„Ù…Ø®Ø²Ù† ÙƒØ§Ù…Ù„" if loc is None else loc,
            "Ø§Ù„Ø¹Ø¯Ø¯ Ø§Ù„ÙØ¹Ù„ÙŠ": d["count"],
            "Ø¹Ø¯Ø¯ Ø§Ù„Ù†Ø¸Ø§Ù…": d["sys_qty"],
        })
    basket_df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù†ÙˆØ¹", "Ø§Ù„Ù…ÙˆÙ‚Ø¹", "Ø§Ù„Ø¹Ø¯Ø¯ Ø§Ù„ÙØ¹Ù„ÙŠ", "Ø¹Ø¯Ø¯ Ø§Ù„Ù†Ø¸Ø§Ù…"])
    table_height = min(800, 80 + len(basket_df) * 35)

    st.dataframe(
        basket_df.sort_values(["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]),
        use_container_width=True,
        height=table_height
    )

    total_count = basket_df["Ø§Ù„Ø¹Ø¯Ø¯ Ø§Ù„ÙØ¹Ù„ÙŠ"].sum() if not basket_df.empty else 0
    st.markdown(f"**Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ù‚Ø·Ø¹ ÙÙŠ Ø§Ù„Ø¬Ø±Ø¯:** {total_count:,}")

    col_clear, col_apply = st.columns(2)
    with col_clear:
        if st.button("ØªÙØ±ÙŠØº Ø§Ù„Ø³Ù„Ø©"):
            st.session_state.stocktake["items"].clear()
            st.toast("ğŸ—‘ï¸ ØªÙ… ØªÙØ±ÙŠØº Ø§Ù„Ø³Ù„Ø©.", icon="ğŸ—‘ï¸", duration=4)

    with col_apply:
        if st.button("ØªØ·Ø¨ÙŠÙ‚ Ø§Ù„ØªØ³ÙˆÙŠØ©"):
            if not st.session_state.stocktake["items"]:
                st.toast("âš ï¸ Ø§Ù„Ø³Ù„Ø© ÙØ§Ø±ØºØ©.", icon="âš ï¸", duration=4)
                return
            # ğŸŸ¦ ØªØ¬Ù…ÙŠØ¹ Ø¬Ù…ÙŠØ¹ Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹ ÙÙŠ Ù…Ù„Ù Excel Ù†Ù‡Ø§Ø¦ÙŠ
            final_export = {}

            # Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹ Ø§Ù„ØªÙŠ ØªÙ… Ø­ÙØ¸Ù‡Ø§ Ù…Ø³Ø¨Ù‚Ø§Ù‹
            for site, df_site in st.session_state.stocktake_sites.items():
                final_export[site] = df_site.copy()

            # Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ø§Ù„Ø­Ø§Ù„ÙŠ Ø¥Ù† ÙƒØ§Ù† Ù…ÙƒØªÙ…Ù„Ø§Ù‹
            if st.session_state.stocktake["scope"] == "loc":
                cur = st.session_state.stocktake["loc"].strip()
                df_current = pd.DataFrame([
                    {
                        "Ø§Ù„ÙƒÙˆØ¯": code,
                        "Ø§Ù„Ù…ÙˆÙ‚Ø¹": cur,
                        "Ø§Ù„Ø¹Ø¯Ø¯ Ø§Ù„ÙØ¹Ù„ÙŠ": d["count"],
                        "Ø¹Ø¯Ø¯ Ø§Ù„Ù†Ø¸Ø§Ù…": d["sys_qty"],
                    }
                    for (code, loc), d in st.session_state.stocktake["items"].items()
                    if loc == cur
                ])
                if not df_current.empty:
                    final_export[cur] = df_current.copy()

            # Ø²Ø± ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ù…Ù„Ù Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠ
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
                for site, df_site in final_export.items():
                    writer_sheet = site[:31]  # Ø§Ø³Ù… Ø§Ù„ÙˆØ±Ù‚Ø© Ø¨Ø­Ø¯ Ø£Ù‚ØµÙ‰ 31 Ø­Ø±Ù
                    df_site.to_excel(writer, index=False, sheet_name=writer_sheet)

            st.download_button(
                "ğŸ“¥ ØªØ­Ù…ÙŠÙ„ Ù…Ù„Ù Ø§Ù„Ø¬Ø±Ø¯ (ÙˆØ±Ù‚Ø© Ù„ÙƒÙ„ Ù…ÙˆÙ‚Ø¹)",
                data=out.getvalue(),
                file_name=f"Ø¬Ø±Ø¯_Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹_{_ts()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            try:
                stock_cur, minlvl_cur, tx_cur, _ = read_all()
                DEFAULT_LOC = "MAIN"
                adj = 0
                for (code, loc), data in st.session_state.stocktake["items"].items():
                    actual, sys_qty = data["count"], data["sys_qty"]
                    delta = actual - sys_qty
                    if delta == 0:
                        continue
                    target_loc = loc or DEFAULT_LOC
                    stock_cur, _ = add_qty(stock_cur, code, target_loc, delta)
                    tx_cur = append_txn(
                        tx_cur, "ADJUST", code, get_part_desc(stock_cur, code),
                        abs(delta),
                        target_loc if delta < 0 else None,
                        target_loc if delta > 0 else None,
                        "STOCKTAKE", "ØªØ³ÙˆÙŠØ© Ø¬Ø±Ø¯")
                    adj += 1
                write_all_with_retry(stock_cur, minlvl_cur, tx_cur)
                st.cache_data.clear()
                st.toast(f"âœ… ØªÙ…Øª Ø§Ù„ØªØ³ÙˆÙŠØ© Ù„Ø¹Ø¯Ø¯ {adj} Ù‚Ø·Ø¹.", icon="âœ…", duration=4)
            except Exception as e:
                st.toast(f"âŒ ÙØ´Ù„ Ø§Ù„ØªØ³ÙˆÙŠØ©: {e}", icon="âŒ", duration=4)

# -------------------------------------------------
# Ø¨Ø§Ù‚ÙŠ Ø§Ù„ØµÙØ­Ø§Øª (Ø¨Ø¯ÙˆÙ† ØªØºÙŠÙŠØ± Ø¬ÙˆÙ‡Ø±ÙŠ Ù„Ø£Ù†Ù‡Ø§ ØªØ¹Ù…Ù„ Ø¬ÙŠØ¯Ù‹Ø§)
# -------------------------------------------------
def _exists_pair(stock: pd.DataFrame, code: str, loc: str) -> bool:
    return ((stock["Ø§Ù„ÙƒÙˆØ¯"] == code) & (stock["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] == loc)).any()


def page_add_new_item():
    st.subheader("Ø¥Ø¶Ø§ÙØ© Ù‚Ø·Ø¹Ø© Ø¬Ø¯ÙŠØ¯Ø©")

    file_status_badge()
    cfg = load_config()
    stock, minlvl, tx, _ = read_all()
    with st.form("add_item_form_simple_ordered", clear_on_submit=False):
        col_qty, col_loc = st.columns(2)
        with col_qty:
            qty = st.number_input("Ø§Ù„ÙƒÙ…ÙŠØ©", min_value=0, value=0, step=1, key="add_qty")
        with col_loc:
            loc = st.text_input("Ø§Ù„Ù…ÙˆÙ‚Ø¹", placeholder="Ù…Ø«Ø§Ù„: Ø±Ù-Ø£1", key="add_loc")
        desc = st.text_input("Ø§Ù„ÙˆØµÙ", placeholder="ÙˆØµÙ Ø§Ù„Ù‚Ø·Ø¹Ø©", key="add_desc")
        col_code, col_orig = st.columns([3, 1])
        with col_code:
            raw_code = st.text_input("Ø§Ù„ÙƒÙˆØ¯", placeholder="Ù…Ø«Ø§Ù„: ABC-123 Ø£Ùˆ ABC-123-S", key="add_code")
        with col_orig:
            is_orig = st.checkbox("Ø£ØµÙ„ÙŠØŸ", value=True, key="add_isorig")
        submitted = st.form_submit_button("Ø¥Ø¶Ø§ÙØ© / Ø²ÙŠØ§Ø¯Ø©")
    if submitted:
        try:
            if not raw_code.strip():
                st.error("Ø§Ù„Ø±Ø¬Ø§Ø¡ Ø¥Ø¯Ø®Ø§Ù„ Ø§Ù„ÙƒÙˆØ¯.")
                return
            if not loc.strip():
                st.error("Ø§Ù„Ø±Ø¬Ø§Ø¡ Ø¥Ø¯Ø®Ø§Ù„ Ø§Ù„Ù…ÙˆÙ‚Ø¹.")
                return
            norm_code = apply_suffix_policy(raw_code, cfg, context="add", checkbox_value=is_orig)
            norm_code = _normalize_code_text(norm_code, cfg, context="add")
            loc = loc.strip()
            qty = int(qty)
            stock_cur, minlvl_cur, tx_cur, _ = read_all()
            if _exists_pair(stock_cur, norm_code, loc):
                current = get_qty(stock_cur, norm_code, loc)
                stock_cur, new_qty = add_qty(stock_cur, norm_code, loc, qty)
                if str(desc).strip():
                    mask = (stock_cur["Ø§Ù„ÙƒÙˆØ¯"] == norm_code) & (stock_cur["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] == loc)
                    cur_desc = str(stock_cur.loc[mask, "Ø§Ù„ÙˆØµÙ"].iloc[0]) if mask.any() else ""
                    if (cur_desc is None) or (str(cur_desc).strip() == ""):
                        stock_cur.loc[mask, "Ø§Ù„ÙˆØµÙ"] = desc.strip()
                if qty > 0:
                    tx_cur = append_txn(tx_cur, "RECEIVE", norm_code,
                                        get_part_desc(stock_cur, norm_code) or desc.strip() or norm_code, qty, None,
                                        loc, user="ADD", note="Add-page increment")
                write_all_with_retry(stock_cur, minlvl_cur, tx_cur)
                st.cache_data.clear()
                st.success(f"ØªÙ…Øª Ø§Ù„Ø²ÙŠØ§Ø¯Ø©: {norm_code} @ {loc} | {current} â†’ {new_qty}")
            else:
                new_row = {"Ø§Ù„ÙƒÙˆØ¯": norm_code, "Ø§Ù„ÙˆØµÙ": desc.strip(), "Ø§Ù„Ù…ÙˆÙ‚Ø¹": loc, "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†": int(qty)}
                stock_cur = pd.concat([stock_cur, pd.DataFrame([new_row])], ignore_index=True)
                if qty > 0:
                    tx_cur = append_txn(tx_cur, "RECEIVE", norm_code, desc.strip() or norm_code, qty, None, loc,
                                        user="ADD", note="Add-page create")
                write_all_with_retry(stock_cur, minlvl_cur, tx_cur)
                st.cache_data.clear()
                st.success(f"ØªÙ…Øª Ø§Ù„Ø¥Ø¶Ø§ÙØ©: {norm_code} @ {loc} Ø¨ÙƒÙ…ÙŠØ© {qty}")
            details, summary = _lookup_code(stock_cur, norm_code)
            if not details.empty:
                st.markdown("**Ø§Ù„ÙˆØ¶Ø¹ Ø§Ù„Ø­Ø§Ù„ÙŠ Ù„Ù„Ù‚Ø·Ø¹Ø©:**")
                st.dataframe(details.sort_values("Ø§Ù„Ù…ÙˆÙ‚Ø¹"), use_container_width=True, height=180)
        except Exception as e:
            st.error(f"ÙØ´Ù„ Ø§Ù„Ø¥Ø¶Ø§ÙØ©/Ø§Ù„Ø²ÙŠØ§Ø¯Ø©: {e}")


def _uploaded_sheets(file) -> List[str]:
    file.seek(0)
    xls = pd.ExcelFile(file, engine="openpyxl")
    return xls.sheet_names


def _read_uploaded_stock(file, sheet_name: str) -> pd.DataFrame:
    file.seek(0)
    xls = pd.ExcelFile(file, engine="openpyxl")
    raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    df = _normalize_stock_cols(_detect_grid(raw))
    return _apply_global_code_normalization(df, context="merge")


def _make_diff(base: pd.DataFrame, incoming: pd.DataFrame, mode: str, only_new: bool) -> pd.DataFrame:
    base_key = base.assign(_key=base["Ø§Ù„ÙƒÙˆØ¯"].astype(str) + "||" + base["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].astype(str))
    inc_key = incoming.assign(_key=incoming["Ø§Ù„ÙƒÙˆØ¯"].astype(str) + "||" + incoming["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].astype(str))
    m = base_key.merge(inc_key, on="_key", how="outer", suffixes=("_current", "_new"))
    if only_new:
        m = m[m["Ø§Ù„ÙƒÙˆØ¯_current"].isna()]

    def _final_qty(row):
        cur = _safe_int(row.get("Ø§Ù„Ù…Ø®Ø²ÙˆÙ†_current"), 0)
        new = _safe_int(row.get("Ø§Ù„Ù…Ø®Ø²ÙˆÙ†_new"), 0)
        if pd.isna(row.get("Ø§Ù„ÙƒÙˆØ¯_new")) and pd.isna(row.get("Ø§Ù„Ù…ÙˆÙ‚Ø¹_new")):
            return cur
        return cur + new if mode == "add" else new

    out = pd.DataFrame({
        "Ø§Ù„ÙƒÙˆØ¯": m["Ø§Ù„ÙƒÙˆØ¯_new"].fillna(m["Ø§Ù„ÙƒÙˆØ¯_current"]).astype(str),
        "Ø§Ù„Ù…ÙˆÙ‚Ø¹": m["Ø§Ù„Ù…ÙˆÙ‚Ø¹_new"].fillna(m["Ø§Ù„Ù…ÙˆÙ‚Ø¹_current"]).astype(str),
        "Ø§Ù„ÙˆØµÙ_Ø­Ø§Ù„ÙŠ": m["Ø§Ù„ÙˆØµÙ_current"],
        "Ø§Ù„ÙˆØµÙ_Ø¬Ø¯ÙŠØ¯": m["Ø§Ù„ÙˆØµÙ_new"],
        "ÙƒÙ…ÙŠØ©_Ø­Ø§Ù„ÙŠØ©": m["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†_current"].fillna(0).astype(int),
        "ÙƒÙ…ÙŠØ©_Ù‚Ø§Ø¯Ù…Ø©": m["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†_new"].fillna(0).astype(int),
    })
    out["Ø§Ù„ÙƒÙ…ÙŠØ©_Ø¨Ø¹Ø¯_Ø§Ù„Ø¯Ù…Ø¬"] = m.apply(_final_qty, axis=1).astype(int)

    def _action(row):
        if row["ÙƒÙ…ÙŠØ©_Ø­Ø§Ù„ÙŠØ©"] == row["Ø§Ù„ÙƒÙ…ÙŠØ©_Ø¨Ø¹Ø¯_Ø§Ù„Ø¯Ù…Ø¬"]:
            return "Ø¨Ø¯ÙˆÙ† ØªØºÙŠÙŠØ±"
        if row["ÙƒÙ…ÙŠØ©_Ø­Ø§Ù„ÙŠØ©"] == 0 and row["ÙƒÙ…ÙŠØ©_Ù‚Ø§Ø¯Ù…Ø©"] > 0 and (
                pd.isna(row["Ø§Ù„ÙˆØµÙ_Ø­Ø§Ù„ÙŠ"]) or str(row["Ø§Ù„ÙˆØµÙ_Ø­Ø§Ù„ÙŠ"]).strip() == ""):
            return "Ø¥Ø¶Ø§ÙØ© ØµÙ Ø¬Ø¯ÙŠØ¯"
        return "ØªØ­Ø¯ÙŠØ« ÙƒÙ…ÙŠØ©"

    out["Ø§Ù„Ø¥Ø¬Ø±Ø§Ø¡"] = out.apply(_action, axis=1)
    return out.sort_values(["Ø§Ù„Ø¥Ø¬Ø±Ø§Ø¡", "Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]).reset_index(drop=True)


def _apply_merge(base: pd.DataFrame, incoming: pd.DataFrame, mode: str,
                 desc_policy: str, only_new: bool) -> Tuple[pd.DataFrame, int, int]:
    updated, added = 0, 0
    result = base.copy()
    if only_new:
        mask = ~incoming.set_index(["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]).index.isin(
            base.set_index(["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]).index
        )
        incoming = incoming[mask].copy()
    for _, r in incoming.iterrows():
        code = str(r["Ø§Ù„ÙƒÙˆØ¯"]).strip()
        loc = str(r["Ø§Ù„Ù…ÙˆÙ‚Ø¹"]).strip()
        qty_new = int(r["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"])
        desc_new = str(r.get("Ø§Ù„ÙˆØµÙ", "")).strip()
        if mode == "add":
            cur = get_qty(result, code, loc)
            result, new_qty = add_qty(result, code, loc, qty_new)
            if (code, loc) in set(base.set_index(["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]).index):
                if new_qty != cur:
                    updated += 1
            else:
                added += 1
        else:
            existed = ((result["Ø§Ù„ÙƒÙˆØ¯"] == code) & (result["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] == loc)).any()
            result = set_qty(result, code, loc, qty_new)
            if existed:
                updated += 1
            else:
                added += 1
        mask = (result["Ø§Ù„ÙƒÙˆØ¯"] == code) & (result["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] == loc)
        cur_desc = str(result.loc[mask, "Ø§Ù„ÙˆØµÙ"].iloc[0]) if mask.any() else ""
        if desc_policy == "replace":
            result.loc[mask, "Ø§Ù„ÙˆØµÙ"] = desc_new
        elif desc_policy == "fill_blank":
            if (cur_desc is None) or (str(cur_desc).strip() == ""):
                result.loc[mask, "Ø§Ù„ÙˆØµÙ"] = desc_new
    return result, updated, added


def page_merge():
    st.subheader("Ø¯Ù…Ø¬ Ù…Ù„Ù Ø¬Ø¯ÙŠØ¯ Ù…Ø¹ Ø§Ù„Ù…Ù„Ù Ø§Ù„Ø­Ø§Ù„ÙŠ")

    file_status_badge()
    base_stock, minlvl, tx, _ = read_all()
    st.caption(f"Ø§Ù„Ø£ÙƒÙˆØ§Ø¯ Ø§Ù„Ø­Ø§Ù„ÙŠØ©: {base_stock['Ø§Ù„ÙƒÙˆØ¯'].nunique():,} | Ø§Ù„ØµÙÙˆÙ: {len(base_stock):,}")
    up = st.file_uploader("Ø§Ø®ØªØ± Ù…Ù„Ù Excel Ø§Ù„ÙŠÙˆÙ…ÙŠ Ù„Ù„Ù‚Ø·Ø¹ Ø§Ù„Ø¬Ø¯ÙŠØ¯Ø©", type=["xlsx", "xls"])
    if not up:
        st.info("Ø§Ø±ÙØ¹ Ù…Ù„ÙÙ‹Ø§ Ù„Ù„Ø¨Ø¯Ø¡.")
        return
    try:
        sheets = _uploaded_sheets(up)
    except Exception as e:
        st.error(f"ØªØ¹Ø°Ù‘Ø± Ù‚Ø±Ø§Ø¡Ø© Ø§Ù„Ù…Ù„Ù: {e}")
        return
    sheet = st.selectbox("Ø§Ø®ØªØ± Ø§Ù„ÙˆØ±Ù‚Ø© Ø¯Ø§Ø®Ù„ Ø§Ù„Ù…Ù„Ù:", options=sheets)
    try:
        incoming = _read_uploaded_stock(up, sheet)
    except Exception as e:
        st.error(f"ÙØ´Ù„ Ø§Ù„ØªØ¹Ø±Ù Ø¹Ù„Ù‰ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø¯Ø§Ø®Ù„ Ø§Ù„ÙˆØ±Ù‚Ø© Ø§Ù„Ù…Ø­Ø¯Ø¯Ø©: {e}")
        return
    cfg = load_config()
    if not incoming.empty:
        incoming["Ø§Ù„ÙƒÙˆØ¯"] = incoming["Ø§Ù„ÙƒÙˆØ¯"].apply(lambda s: _normalize_code_text(s, cfg, context="merge"))
    st.success(f"ØªÙ… Ø§Ù„ØªØ¹Ø±Ù Ø¹Ù„Ù‰ {len(incoming)} ØµÙÙ‹Ø§ Ù…Ù† Ø§Ù„Ù…Ù„Ù Ø§Ù„Ø¬Ø¯ÙŠØ¯. Ù…Ø¹Ø§ÙŠÙ†Ø©:")
    st.dataframe(incoming.head(30), use_container_width=True, height=240)
    st.markdown("### Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ø¯Ù…Ø¬")
    c1, c2, c3 = st.columns(3)
    with c1:
        mode = st.radio("Ø¥Ø³ØªØ±Ø§ØªÙŠØ¬ÙŠØ© Ø§Ù„ÙƒÙ…ÙŠØ©", ["Ø§Ø³ØªØ¨Ø¯Ø§Ù„ Ø§Ù„ÙƒÙ…ÙŠØ© (Set)", "Ø¥Ø¶Ø§ÙØ© Ø¹Ù„Ù‰ Ø§Ù„ÙƒÙ…ÙŠØ© (Add)"], horizontal=False)
        mode_key = "set" if mode.startswith("Ø§Ø³ØªØ¨Ø¯Ø§Ù„") else "add"
    with c2:
        desc_policy = st.selectbox("Ø³ÙŠØ§Ø³Ø© Ø§Ù„ÙˆØµÙ", ["Ù„Ø§ ØªØºÙŠÙ‘Ø± Ø§Ù„ÙˆØµÙ Ø§Ù„Ø­Ø§Ù„ÙŠ", "Ø­Ø¯Ù‘Ø« Ø§Ù„ÙˆØµÙ Ø¥Ø°Ø§ ÙƒØ§Ù† Ø§Ù„Ø­Ø§Ù„ÙŠ ÙØ§Ø±ØºÙ‹Ø§",
                                                   "Ø§Ø³ØªØ¨Ø¯Ù„ Ø§Ù„ÙˆØµÙ Ø¯Ø§Ø¦Ù…Ù‹Ø§ Ø¨Ø§Ù„Ù‚Ø§Ø¯Ù…"])
        desc_key = {"Ù„Ø§ ØªØºÙŠÙ‘Ø± Ø§Ù„ÙˆØµÙ Ø§Ù„Ø­Ø§Ù„ÙŠ": "keep", "Ø­Ø¯Ù‘Ø« Ø§Ù„ÙˆØµÙ Ø¥Ø°Ø§ ÙƒØ§Ù† Ø§Ù„Ø­Ø§Ù„ÙŠ ÙØ§Ø±ØºÙ‹Ø§": "fill_blank",
                    "Ø§Ø³ØªØ¨Ø¯Ù„ Ø§Ù„ÙˆØµÙ Ø¯Ø§Ø¦Ù…Ù‹Ø§ Ø¨Ø§Ù„Ù‚Ø§Ø¯Ù…": "replace"}[desc_policy]
    with c3:
        only_new = st.checkbox("Ø§Ø³ØªÙŠØ±Ø§Ø¯ Ø§Ù„Ø£ÙƒÙˆØ§Ø¯/Ø§Ù„Ù…ÙˆØ§Ù‚Ø¹ Ø§Ù„Ø¬Ø¯ÙŠØ¯Ø© ÙÙ‚Ø·", value=False)
    st.markdown("### Ø§Ù„Ù…Ø¹Ø§ÙŠÙ†Ø© Ù‚Ø¨Ù„ Ø§Ù„Ø­ÙØ¸ (Diff)")
    diff_df = _make_diff(base_stock, incoming, mode_key, only_new)
    st.dataframe(diff_df, use_container_width=True, height=320)
    add_count = (diff_df["Ø§Ù„Ø¥Ø¬Ø±Ø§Ø¡"] == "Ø¥Ø¶Ø§ÙØ© ØµÙ Ø¬Ø¯ÙŠØ¯").sum()
    upd_count = (diff_df["Ø§Ù„Ø¥Ø¬Ø±Ø§Ø¡"] == "ØªØ­Ø¯ÙŠØ« ÙƒÙ…ÙŠØ©").sum()
    nochg_count = (diff_df["Ø§Ù„Ø¥Ø¬Ø±Ø§Ø¡"] == "Ø¨Ø¯ÙˆÙ† ØªØºÙŠÙŠØ±").sum()
    st.caption(f"Ø¥Ø­ØµØ§Ø¦ÙŠØ§Øª: Ø¬Ø¯ÙŠØ¯: {add_count} | ØªØ­Ø¯ÙŠØ«: {upd_count} | Ø¨Ø¯ÙˆÙ† ØªØºÙŠÙŠØ±: {nochg_count}")
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as w:
        diff_df.to_excel(w, index=False, sheet_name="Diff")
        incoming.to_excel(w, index=False, sheet_name="Incoming")
        base_stock.to_excel(w, index=False, sheet_name="Current")
    st.download_button("ØªÙ†Ø²ÙŠÙ„ ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ù…Ù‚Ø§Ø±Ù†Ø© (Excel)", data=out.getvalue(),
                       file_name="ØªÙ‚Ø±ÙŠØ±_Ø¯Ù…Ø¬_Ø§Ù„Ù…Ø®Ø²ÙˆÙ†.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if st.button("ØªÙ†ÙÙŠØ° Ø§Ù„Ø¯Ù…Ø¬ ÙˆØ§Ù„Ø­ÙØ¸ Ø¯Ø§Ø®Ù„ Ø§Ù„Ù…Ù„Ù Ø§Ù„Ø­Ø§Ù„ÙŠ"):
        try:
            merged, updated, added = _apply_merge(base_stock, incoming, mode_key, desc_key, only_new)
            tx = append_txn(tx, "ADJUST", "BULK_MERGE", "Ø¯Ù…Ø¬ Ù…Ù„Ù ÙŠÙˆÙ…ÙŠ", int(len(incoming)), None, None, user="MERGE",
                            note=f"mode={mode_key}, desc={desc_key}, only_new={only_new}")
            write_all_with_retry(merged, minlvl, tx)
            st.cache_data.clear()
            st.success(f"ØªÙ… Ø§Ù„Ø¯Ù…Ø¬ Ø¨Ù†Ø¬Ø§Ø­. ØªÙ…Øª Ø¥Ø¶Ø§ÙØ© {added} ÙˆØªØ­Ø¯ÙŠØ« {updated} ØµÙÙ‹Ø§.")
            if st.button("Ø§Ù„Ø§Ù†ØªÙ‚Ø§Ù„ Ø¥Ù„Ù‰ Ù„ÙˆØ­Ø© Ø§Ù„ØªØ­ÙƒÙ…"):
                nav_to("Ù„ÙˆØ­Ø© Ø§Ù„ØªØ­ÙƒÙ…")
        except Exception as e:
            st.error(f"ÙØ´Ù„ Ø§Ù„Ø¯Ù…Ø¬: {e}")


def page_data_editor():
    st.subheader("ØªØ­Ø±ÙŠØ± Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ø¨Ø§Ø´Ø±Ø© (Stock)")

    file_status_badge()
    stock, minlvl, tx, _ = read_all()
    edited_stock = st.data_editor(
        stock,
        key="stock_editor",
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "Ø§Ù„ÙƒÙˆØ¯": st.column_config.TextColumn(required=True),
            "Ø§Ù„ÙˆØµÙ": st.column_config.TextColumn(required=False),
            "Ø§Ù„Ù…ÙˆÙ‚Ø¹": st.column_config.TextColumn(required=True),
            "Ø§Ù„Ù…Ø®Ø²ÙˆÙ†": st.column_config.NumberColumn(min_value=0, step=1),
        },
    )
    if st.button("Ø­ÙØ¸ Ø§Ù„ØªØºÙŠÙŠØ±Ø§Øª"):
        try:
            cfg = load_config()
            edited_stock["Ø§Ù„ÙƒÙˆØ¯"] = edited_stock["Ø§Ù„ÙƒÙˆØ¯"].apply(
                lambda s: _normalize_code_text(s, cfg, context="editor"))
            edited_stock["Ø§Ù„Ù…ÙˆÙ‚Ø¹"] = edited_stock["Ø§Ù„Ù…ÙˆÙ‚Ø¹"].fillna("").astype(str).str.strip()
            edited_stock["Ø§Ù„ÙˆØµÙ"] = edited_stock["Ø§Ù„ÙˆØµÙ"].fillna("").astype(str).str.strip()
            edited_stock["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"] = pd.to_numeric(edited_stock["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"], errors="coerce").fillna(0).astype(int)
            write_all_with_retry(edited_stock, minlvl, tx)
            st.cache_data.clear()
            st.success("ØªÙ… Ø§Ù„Ø­ÙØ¸ Ø¯Ø§Ø®Ù„ Ù†ÙØ³ Ø§Ù„Ù…Ù„Ù.")
        except Exception as e:
            st.error(f"Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ù„Ø­ÙØ¸: {e}")


def page_operations():
    st.subheader("Ø§Ù„Ø¹Ù…Ù„ÙŠØ§Øª (ØµØ±Ù / ØªØ­ÙˆÙŠÙ„)")

    file_status_badge()
    stock, minlvl, tx, _ = read_all()
    cfg = load_config()
    min_level = int(cfg.get("global_min_level", 2))
    codes_list = get_unique_codes(stock)
    locs_list = get_unique_locations(stock)
    op = st.selectbox("Ø§Ø®ØªØ± Ø§Ù„Ø¹Ù…Ù„ÙŠØ©", ["ØµØ±Ù (ISSUE)", "ØªØ­ÙˆÙŠÙ„ (TRANSFER)"])
    mode_code = st.radio("Ø·Ø±ÙŠÙ‚Ø© Ø¥Ø¯Ø®Ø§Ù„ Ø§Ù„ÙƒÙˆØ¯", ["ÙƒØªØ§Ø¨ÙŠ", "Ù‚Ø§Ø¦Ù…Ø©"], horizontal=True, index=0)
    mode_loc = st.radio("Ø·Ø±ÙŠÙ‚Ø© Ø¥Ø¯Ø®Ø§Ù„ Ø§Ù„Ù…ÙˆÙ‚Ø¹", ["ÙƒØªØ§Ø¨ÙŠ", "Ù‚Ø§Ø¦Ù…Ø©"], horizontal=True, index=0)

    def input_code(label_key: str):
        if mode_code == "Ù‚Ø§Ø¦Ù…Ø©" and codes_list:
            return st.selectbox(label_key, options=codes_list, key=label_key + "_select"), None
        cols = st.columns([3, 1])
        with cols[0]:
            raw = st.text_input(label_key, key=label_key + "_text", placeholder="Ø§Ù…Ø³Ø­ Ø§Ù„Ø¨Ø§Ø±ÙƒÙˆØ¯ Ø£Ùˆ Ø§ÙƒØªØ¨ Ø§Ù„ÙƒÙˆØ¯")
        with cols[1]:
            orig = st.checkbox("Ø£ØµÙ„ÙŠØŸ", value=True, key=label_key + "_isorig")
        return raw, orig

    def input_loc(label_key: str):
        if mode_loc == "Ù‚Ø§Ø¦Ù…Ø©" and locs_list:
            return st.selectbox(label_key, options=locs_list, key=label_key + "_select")
        return st.text_input(label_key, key=label_key + "_text")

    def preview_qty(code: str, loc: Optional[str] = None):
        if not code:
            return
        details, summary = _lookup_code(stock, code)
        if details.empty:
            st.info("Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯ ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯ ÙÙŠ Ø§Ù„Ù…Ø®Ø²ÙˆÙ†.")
            return
        with st.expander("Ø¹Ø±Ø¶ Ø³Ø±ÙŠØ¹ Ù„Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ Ù„Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯", expanded=False):
            st.dataframe(details.sort_values("Ø§Ù„Ù…ÙˆÙ‚Ø¹"), use_container_width=True, height=160)
            if loc:
                st.caption(f"Ø§Ù„ÙƒÙ…ÙŠØ© ÙÙŠ [{loc}]: {get_qty(stock, code, loc)}")

    with st.form("ops_form"):
        col1, col2 = st.columns(2)
        with col1:
            code_raw, isorig = input_code("Ø§Ù„ÙƒÙˆØ¯")
        with col2:
            if mode_code == "Ù‚Ø§Ø¦Ù…Ø©":
                norm_code = _normalize_code_text(code_raw or "", cfg, context="ops")
            else:
                norm_code = apply_suffix_policy(code_raw or "", cfg, context="ops", checkbox_value=isorig)
            desc_default = get_part_desc(stock, norm_code) if norm_code else ""
            description = st.text_input("Ø§Ù„ÙˆØµÙ (Ø§Ø®ØªÙŠØ§Ø±ÙŠ)", value=desc_default)
        qty = st.number_input("Ø§Ù„ÙƒÙ…ÙŠØ©", min_value=1, value=1, step=1)
        note = st.text_input("Ù…Ù„Ø§Ø­Ø¸Ø©")
        user = st.text_input("Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… (Ø§Ø®ØªÙŠØ§Ø±ÙŠ)")
        if op == "ØµØ±Ù (ISSUE)":
            from_loc = input_loc("Ù…Ù† Ù…ÙˆÙ‚Ø¹")
            if norm_code and from_loc:
                preview_qty(norm_code, from_loc)
            submitted = st.form_submit_button("ØªÙ†ÙÙŠØ° Ø§Ù„ØµØ±Ù")
            if submitted:
                if not norm_code or not from_loc:
                    st.error("Ø£Ø¯Ø®Ù„ Ø§Ù„ÙƒÙˆØ¯ ÙˆÙ…ÙˆÙ‚Ø¹ Ø§Ù„ØµØ±Ù.")
                else:
                    current = get_qty(stock, norm_code, from_loc)
                    if current <= 0:
                        st.markdown(
                            f"<div class='error-box'>âŒ Ø§Ù„ÙƒÙˆØ¯ <b>{norm_code}</b> ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯/ØµÙØ± ÙÙŠ {from_loc}.</div>",
                            unsafe_allow_html=True)
                    elif int(qty) > current:
                        st.markdown(
                            f"<div class='error-box'>âŒ Ø§Ù„ÙƒÙ…ÙŠØ© Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø© ({int(qty)}) Ø£ÙƒØ¨Ø± Ù…Ù† Ø§Ù„Ù…ØªØ§Ø­ ({current}) ÙÙŠ {from_loc}.</div>",
                            unsafe_allow_html=True)
                    else:
                        try:
                            stock, new_qty = add_qty(stock, norm_code, from_loc, -int(qty))
                        except ValueError as e:
                            st.error(str(e))
                        else:
                            tx = append_txn(tx, "ISSUE", norm_code, get_part_desc(stock, norm_code), int(qty), from_loc,
                                            None, user, note)
                            write_all_with_retry(stock, minlvl, tx)
                            st.success(f"ØªÙ… Ø§Ù„ØµØ±Ù. Ø§Ù„ÙƒÙ…ÙŠØ© Ø§Ù„Ø­Ø§Ù„ÙŠØ© ÙÙŠ {from_loc}: {new_qty}")
                            if new_qty == 0:
                                st.error("âš ï¸ Ù†ÙØ¯Øª Ø§Ù„ÙƒÙ…ÙŠØ© Ù„Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯ ÙÙŠ Ù‡Ø°Ø§ Ø§Ù„Ù…ÙˆÙ‚Ø¹.")
                            elif new_qty <= min_level:
                                st.warning(f"âš ï¸ Ø§Ù„ÙƒÙ…ÙŠØ© Ù…Ù†Ø®ÙØ¶Ø© (â‰¤ {min_level}).")
        elif op == "ØªØ­ÙˆÙŠÙ„ (TRANSFER)":
            c1, c2 = st.columns(2)
            with c1:
                from_loc = input_loc("Ù…Ù† Ù…ÙˆÙ‚Ø¹")
            with c2:
                to_loc = input_loc("Ø¥Ù„Ù‰ Ù…ÙˆÙ‚Ø¹")
            if norm_code and from_loc:
                preview_qty(norm_code, from_loc)
            submitted = st.form_submit_button("ØªÙ†ÙÙŠØ° Ø§Ù„ØªØ­ÙˆÙŠÙ„")
            if submitted:
                if not norm_code or not from_loc or not to_loc:
                    st.error("Ø£Ø¯Ø®Ù„ Ø§Ù„ÙƒÙˆØ¯ ÙˆØ§Ù„Ù…ÙˆÙ‚Ø¹ÙŠÙ†.")
                elif from_loc == to_loc:
                    st.error("Ø§Ø®ØªØ± Ù…ÙˆÙ‚Ø¹ÙŠÙ† Ù…Ø®ØªÙ„ÙÙŠÙ†.")
                else:
                    current = get_qty(stock, norm_code, from_loc)
                    if current <= 0:
                        st.error("Ù„Ø§ ØªÙˆØ¬Ø¯ ÙƒÙ…ÙŠØ© Ù„Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯ ÙÙŠ Ù…ÙˆÙ‚Ø¹ Ø§Ù„ØªØ­ÙˆÙŠÙ„ (Ù…Ù†).")
                    elif int(qty) > current:
                        st.error(f"Ø§Ù„ÙƒÙ…ÙŠØ© Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø© ({int(qty)}) Ø£ÙƒØ¨Ø± Ù…Ù† Ø§Ù„Ù…ØªØ§Ø­ ({current}) ÙÙŠ {from_loc}.")
                    else:
                        try:
                            stock, new_from = add_qty(stock, norm_code, from_loc, -int(qty))
                        except ValueError as e:
                            st.error(str(e))
                        else:
                            stock, new_to = add_qty(stock, norm_code, to_loc, int(qty))
                            tx = append_txn(tx, "TRANSFER", norm_code, get_part_desc(stock, norm_code), int(qty),
                                            from_loc, to_loc, user, note)
                            write_all_with_retry(stock, minlvl, tx)
                            st.success(f"ØªÙ… Ø§Ù„ØªØ­ÙˆÙŠÙ„. Ø§Ù„Ù…ØªØ¨Ù‚ÙŠ ÙÙŠ {from_loc}: {new_from} â€” Ø§Ù„Ø­Ø§Ù„ÙŠ ÙÙŠ {to_loc}: {new_to}")
                            if new_from == 0:
                                st.error(f"âš ï¸ Ù†ÙØ¯Øª Ø§Ù„ÙƒÙ…ÙŠØ© ÙÙŠ {from_loc}.")
                            elif new_from <= min_level:
                                st.warning(f"âš ï¸ Ø§Ù„ÙƒÙ…ÙŠØ© Ù…Ù†Ø®ÙØ¶Ø© ÙÙŠ {from_loc} (â‰¤ {min_level}).")


def page_import_export():
    st.subheader("Ø§Ø³ØªÙŠØ±Ø§Ø¯ / ØªØµØ¯ÙŠØ±")

    file_status_badge()
    stock, minlvl, tx, _ = read_all()
    st.markdown("### ØªÙ†Ø²ÙŠÙ„ Ù†Ø³Ø®Ø© Ø¹Ù…Ù„ (Stock + Transactions)")
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as w:
        stock.to_excel(w, index=False, sheet_name="Stock")
        tx.to_excel(w, index=False, sheet_name="Transactions")
    st.download_button(
        "ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ù…Ø®Ø²ÙˆÙ†_Ø§Ù„Ø­Ø§Ù„ÙŠ.xlsx",
        data=out.getvalue(),
        file_name="Ø§Ù„Ù…Ø®Ø²ÙˆÙ†_Ø§Ù„Ø­Ø§Ù„ÙŠ.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.markdown("---")
    st.caption("Ù„Ø¹Ù…Ù„ÙŠØ© Ø¯Ù…Ø¬ Ù…ØªÙ‚Ø¯Ù…Ø© Ø§Ø³ØªØ®Ø¯Ù… ØµÙØ­Ø© 'Ø¯Ù…Ø¬ Ù…Ù„Ù Ø¬Ø¯ÙŠØ¯' Ù…Ù† Ø§Ù„Ù‚Ø§Ø¦Ù…Ø©.")


def page_settings():
    st.subheader("Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª")

    st.caption(f"Ø§Ù„Ù…Ø³Ø§Ø± Ø§Ù„Ø­Ø§Ù„ÙŠ Ù„Ù…Ù„Ù Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª: {EXCEL_PATH}")
    file_status_badge()
    cfg = load_config()
    colA, colB = st.columns([2, 2])
    with colA:
        min_level = st.number_input("Ø§Ù„Ø­Ø¯ Ø§Ù„Ø£Ø¯Ù†Ù‰ Ø§Ù„Ø§ÙØªØ±Ø§Ø¶ÙŠ Ù„Ù„ØªÙ†Ø¨ÙŠÙ‡ (Ø¥Ø¹Ø§Ø¯Ø© Ø§Ù„Ø·Ù„Ø¨)", min_value=0,
                                    value=int(cfg.get("global_min_level", 2)), step=1)
        code_case = st.selectbox("ØªØ·Ø¨ÙŠØ¹ Ø­Ø±ÙˆÙ Ø§Ù„ÙƒÙˆØ¯", ["upper", "lower", "none"],
                                 index=["upper", "lower", "none"].index(cfg.get("code_case", "upper")))
    with colB:
        auto_suffix_mode = st.selectbox("Ù…Ù†Ø·Ù‚ Ø§Ù„Ù„Ø§Ø­Ù‚Ø© -S (ØªÙ…ÙŠÙŠØ² Ø§Ù„Ø£ØµÙ„ÙŠ)", ["by_checkbox", "always", "off"],
                                        index=["by_checkbox", "always", "off"].index(
                                            cfg.get("auto_suffix_mode", "by_checkbox")))
        suffix_text = st.text_input("Ù†Øµ Ø§Ù„Ù„Ø§Ø­Ù‚Ø© Ù„Ù„Ø£ØµÙ„ÙŠ", value=str(cfg.get("suffix_text", "-S")))
        apply_on = st.multiselect("ØªØ·Ø¨ÙŠÙ‚ Ø§Ù„ØªØ·Ø¨ÙŠØ¹ Ø§Ù„Ø­Ø±ÙÙŠ Ø¹Ù†Ø¯",
                                  options=["scan", "bulk", "merge", "ops", "editor", "import", "add"],
                                  default=_unique_order(cfg.get("suffix_apply_on",
                                                                ["scan", "bulk", "merge", "ops", "editor", "import",
                                                                 "add"])))
    contexts_all = ["scan", "bulk", "ops", "stocktake", "merge", "editor", "import", "add"]
    suffix_ctx = st.multiselect(
        "ØªØ·Ø¨ÙŠÙ‚ Ù…Ù†Ø·Ù‚ Ø§Ù„Ù„Ø§Ø­Ù‚Ø© -S ÙÙŠ Ù‡Ø°Ù‡ Ø§Ù„Ø³ÙŠØ§Ù‚Ø§Øª",
        options=contexts_all,
        default=_unique_order(cfg.get("suffix_apply_on_contexts", ["scan", "bulk", "ops", "stocktake", "add"]))
    )
    if st.button("Ø­ÙØ¸ Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª"):
        cfg["global_min_level"] = int(min_level)
        cfg["code_case"] = code_case
        cfg["auto_suffix_mode"] = auto_suffix_mode
        cfg["suffix_text"] = suffix_text
        cfg["suffix_apply_on"] = apply_on
        cfg["suffix_apply_on_contexts"] = suffix_ctx
        cfg["enable_backups"] = False
        cfg["backup_keep"] = 0
        save_config(cfg)
        st.success("ØªÙ… Ø­ÙØ¸ Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª.")
        st.cache_data.clear()
    try:
        _, _, _, names = read_all()
        st.write("Ø§Ù„Ø£ÙˆØ±Ø§Ù‚ Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© Ø§Ù„Ø¢Ù†:", names)
    except Exception:
        pass
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if st.button("ØªØ­Ø¯ÙŠØ« Ø§Ù„ÙƒØ§Ø´ / Ø¥Ø¹Ø§Ø¯Ø© Ø§Ù„ØªØ­Ù…ÙŠÙ„"):
            st.cache_data.clear()
            st.success("ØªÙ… Ù…Ø³Ø­ Ø§Ù„ÙƒØ§Ø´. Ø£Ø¹Ø¯ ØªØ­Ù…ÙŠÙ„ Ø§Ù„ØµÙØ­Ø© (Ctrl+F5).")
    with c2:
        if st.button("Ø­Ø°Ù ÙˆØ±Ù‚Ø© MinLevels (Ø¥Ù† ÙˆÙØ¬Ø¯Øª)"):
            _drop_sheet_if_exists(EXCEL_PATH, "MinLevels")
            st.cache_data.clear()
            st.success("ØªÙ… Ø­Ø°Ù ÙˆØ±Ù‚Ø© MinLevels.")
    with c3:
        if st.button("ØªØ¬Ù…ÙŠØ¹ Ø§Ù„Ù…ÙƒØ±Ø±Ø§Øª ÙˆØ­ÙØ¸"):
            stock, minlvl, tx, _ = read_all()
            stock2 = _compact_stock(stock)
            write_all_with_retry(stock2, minlvl, tx)
            st.cache_data.clear()
            st.success("ØªÙ… Ø§Ù„ØªØ¬Ù…ÙŠØ¹ ÙˆØ§Ù„Ø­ÙØ¸.")
    with c4:
        if st.button("Ø¥Ù†Ø´Ø§Ø¡/ØªØ¬Ø¯ÙŠØ¯ Ø§Ù„Ù‡ÙŠÙƒÙ„ Ø§Ù„Ù‚ÙŠØ§Ø³ÙŠ"):
            ensure_excel_file()
            st.success("ØªÙ… Ø§Ù„ØªØ£ÙƒØ¯ Ù…Ù† ÙˆØ¬ÙˆØ¯ Ø§Ù„Ù…Ù„Ù ÙˆØ§Ù„Ø£ÙˆØ±Ø§Ù‚ Ø§Ù„Ù‚ÙŠØ§Ø³ÙŠØ© (Ø¨Ø¯ÙˆÙ† MinLevels).")


def render_credits():
    year = datetime.now().year
    with st.sidebar:
        st.markdown(f"<div class='sidebar-credit'>Â© {year} â€” <b>{DEV_NAME}</b></div>", unsafe_allow_html=True)
    st.markdown(f"<div class='dev-credit'>Â© {year} â€” <b>{DEV_NAME}</b></div>", unsafe_allow_html=True)


def page_dashboard():
    st.subheader("Ù„ÙˆØ­Ø© Ø§Ù„ØªØ­ÙƒÙ…")

    file_status_badge()
    stock, minlvl, tx, _ = read_all()
    cfg = load_config()
    min_level = int(cfg.get("global_min_level", 2))
    total_items = stock["Ø§Ù„ÙƒÙˆØ¯"].nunique()
    total_qty = int(stock["Ø§Ù„Ù…Ø®Ø²ÙˆÙ†"].sum()) if not stock.empty else 0
    loc_count = len(get_unique_locations(stock))
    suf = _suffix_to_use(cfg)
    # âœ… Ø§Ù„Ø¢Ù†: Ø§Ù„Ø£ØµÙ„ÙŠ = Ù„Ø§ ÙŠØ­ØªÙˆÙŠ Ø¹Ù„Ù‰ -S
    orig_count = (~stock["Ø§Ù„ÙƒÙˆØ¯"].astype(str).str.endswith(suf)).sum()
    comm_count = total_items - orig_count
    low_df, oos_df = compute_low_and_oos(stock, min_level)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Ø¹Ø¯Ø¯ Ø§Ù„Ø£ÙƒÙˆØ§Ø¯", total_items)
    c2.metric("Ø£ÙƒÙˆØ§Ø¯ Ø£ØµÙ„ÙŠØ©", int(orig_count))
    c3.metric("Ø£ÙƒÙˆØ§Ø¯ ØªØ¬Ø§Ø±ÙŠØ©", int(comm_count))
    c4.metric("ØºÙŠØ± Ù…ØªÙˆÙØ± (0)", int(len(oos_df)))
    c5.metric(f"Ù‚Ø±ÙŠØ¨ Ù…Ù† Ø§Ù„Ù†ÙØ§Ø¯ (â‰¤ {min_level})", int(len(low_df)))
    if len(oos_df) > 0:
        st.error("Ù‚Ø·Ø¹ ØºÙŠØ± Ù…ØªÙˆÙØ±Ø© Ø­Ø§Ù„ÙŠÙ‹Ø§ (Ø§Ù„Ù…Ø®Ø²ÙˆÙ† = 0):")
        st.dataframe(oos_df, use_container_width=True, height=200)
    if len(low_df) > 0:
        st.warning(f"Ù‚Ø·Ø¹ Ø§Ù‚ØªØ±Ø¨Øª Ù…Ù† Ø§Ù„Ù†ÙØ§Ø¯ (â‰¤ {min_level}):")
        st.dataframe(low_df, use_container_width=True, height=200)
    st.markdown("### Ø§Ù„Ù…Ø®Ø²ÙˆÙ† Ø§Ù„Ø­Ø§Ù„ÙŠ")
    st.dataframe(stock.sort_values(["Ø§Ù„ÙƒÙˆØ¯", "Ø§Ù„Ù…ÙˆÙ‚Ø¹"]), use_container_width=True, height=420)


# -------------------------------------------------
# Main
# -------------------------------------------------
def main():
    st.title("Ù†Ø¸Ø§Ù… Ø¥Ø¯Ø§Ø±Ø© Ù…Ø®Ø²ÙˆÙ† Ù‚Ø·Ø¹ Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª (ÙŠØ¹ØªÙ…Ø¯ Ù…Ù„Ù Excel ÙˆØ§Ø­Ø¯)")
    st.caption("Ù‚Ø±Ø§Ø¡Ø© ÙˆÙƒØªØ§Ø¨Ø© Ù…Ø¨Ø§Ø´Ø±Ø© Ø¯Ø§Ø®Ù„: " + EXCEL_PATH)
    if "menu" not in st.session_state:
        st.session_state.menu = "Ù„ÙˆØ­Ø© Ø§Ù„ØªØ­ÙƒÙ…"
    default_index = PAGES.index(st.session_state.menu) if st.session_state.menu in PAGES else 0
    menu = st.sidebar.radio("Ø§Ù„Ù‚Ø§Ø¦Ù…Ø©", PAGES, index=default_index, key="sidebar_menu")
    if menu != st.session_state.menu:
        st.session_state.menu = menu
    if menu == "Ù„ÙˆØ­Ø© Ø§Ù„ØªØ­ÙƒÙ…":
        page_dashboard()
    elif menu == "Ø¨Ø­Ø«/Ù…Ø³Ø­":
        page_find_and_scan()
    elif menu == "Ø§Ù„Ø¹Ù…Ù„ÙŠØ§Øª":
        page_operations()
    elif menu == "Ø§Ù„Ø¬Ø±Ø¯":
        page_stocktake()
    elif menu == "Ø¥Ø¶Ø§ÙØ© Ù‚Ø·Ø¹Ø© Ø¬Ø¯ÙŠØ¯Ø©":
        page_add_new_item()
    elif menu == "Ø¯Ù…Ø¬ Ù…Ù„Ù Ø¬Ø¯ÙŠØ¯":
        page_merge()
    elif menu == "ØªØ­Ø±ÙŠØ± Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª":
        page_data_editor()
    elif menu == "Ø§Ø³ØªÙŠØ±Ø§Ø¯/ØªØµØ¯ÙŠØ±":
        page_import_export()
    elif menu == "Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª":
        page_settings()
    render_credits()


if __name__ == "__main__":
    main()
